#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import io
import os
import time
import re
import csv
import json
import sqlite3
import struct
import zipfile
import hashlib
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, jsonify, send_file, request, session as flask_session, redirect, url_for
from functools import wraps
import requests
import secrets
import threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF

# --- IMPORTATION DU MODULE CARNET DE PLAINE ---
from interventions import interventions_bp, init_db

DASHBOARD_VERSION = "15.6"

app = Flask(__name__)

# Clé de session : générée aléatoirement une seule fois, puis persistée sur disque (même
# principe que password_override.txt) -- sans ça, un redémarrage du serveur invalide
# silencieusement TOUTES les sessions ouvertes (l'utilisateur reste "connecté" côté
# navigateur mais chaque requête API échoue en 401, sans message clair sur la vraie cause).
# Particulièrement génant en développement/test, où le serveur redémarre souvent.
try:
    with open("secret_key.txt", "r") as _f:
        _secret = _f.read().strip()
        if not _secret or len(_secret) < 32:
            raise ValueError("clé absente ou trop courte")
except Exception:
    _secret = secrets.token_hex(32)
    try:
        with open("secret_key.txt", "w") as _f:
            _f.write(_secret)
    except Exception:
        pass  # pas bloquant : la cle reste generee en memoire pour cette execution
app.secret_key = _secret

# Aucun endpoit ne reçoit de fichier uploadé côté serveur (les imports KML/GeoJSON/CSV sont
# entièrement parsés côté navigateur, seules des données déjà structurées sont envoyées à
# l'API) -- une requête entrante ne devrait donc jamais dépasser quelques centaines de Ko en
# usage normal, même pour un import de sous-parcelles avec des contours très détaillés. Sans
# limite, une requête anormalement volumineuse (bug client ou tentative malveillante) pouvait
# consommer une quantité de mémoire arbitraire. 10 Mo laisse une marge confortable.
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024

# Sécurité du cookie de session : HttpOnly (inaccessible en JavaScript, protège même en cas
# de faille XSS) et SameSite=Lax (limite l'envoi du cookie sur des requêtes cross-site,
# protection de base contre le CSRF) sont déjà les valeurs par défaut de Flask -- explicitées
# ici pour ne jamais en dépendre implicitement. "Secure" (cookie envoyé uniquement en HTTPS)
# n'est PAS activé par défaut : la majorité des installations tournent en HTTP simple sur un
# réseau local, et l'activer sans condition les aurait bloquées hors de tout accès. Réglable
# via config.json -> "security": {"force_https_cookie": true} pour qui dispose d'un
# reverse-proxy HTTPS devant l'application (voir Notice, section Maintenance).
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
try:
    with open("config.json") as _f:
        app.config['SESSION_COOKIE_SECURE'] = bool(json.load(_f).get("security", {}).get("force_https_cookie", False))
except Exception:
    app.config['SESSION_COOKIE_SECURE'] = False

# ================= AUTHENTIFICATION =================
LOGIN_USER = "admin"       # ← Changer ici
_DEFAULT_LOGIN_PASSWORD = "admin"  # ← Changer ici (via l'interface, pas ce fichier).
# Identifiant/mot de passe volontairement simples ("admin"/"admin") pour un premier acces
# immediat sans etape supplementaire -- le bandeau d'alerte "mot de passe par defaut"
# (voir plus bas, comparaison avec _DEFAULT_LOGIN_PASSWORD) rappelle de le changer des
# que le dashboard est expose au-dela du seul reseau local.


def _hash_password(password, salt=None):
    """
    Hache un mot de passe avec PBKDF2-HMAC-SHA256 (100 000 itérations) et un sel aléatoire
    propre à ce mot de passe -- contrairement à un simple SHA-256 non salé (vulnérable aux
    tables arc-en-ciel), ou pire, à un stockage en clair (ce que faisait ce fichier jusqu'ici
    malgré une documentation qui prétendait déjà le contraire).
    Retourne "sel_hex$empreinte_hex".
    """
    if salt is None:
        salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), bytes.fromhex(salt), 100_000).hex()
    return f"{salt}${digest}"


def _verify_password(password, stored):
    """
    Vérifie un mot de passe contre sa forme stockée. Reconnaît aussi l'ancien format en
    clair (sans '$') pour ne pas invalider un mot de passe déjà en place lors de la mise à
    jour -- il sera transparently re-haché à la prochaine connexion réussie.
    """
    if '$' not in stored:
        return password == stored
    salt, digest = stored.split('$', 1)
    try:
        check = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), bytes.fromhex(salt), 100_000).hex()
    except Exception:
        return False
    return check == digest


def _migrate_legacy_plaintext_password(plain_password):
    """
    Si le mot de passe stocké est encore dans l'ancien format en clair (sans '$'), le
    re-hache et le persiste au premier login réussi -- migration transparente, sans action
    requise de la part de l'utilisateur. Le mot de passe par défaut d'usine (jamais changé)
    n'est volontairement PAS migré : il reste comparable tel quel pour que le bandeau
    "mot de passe par défaut" continue de le détecter simplement.
    """
    global LOGIN_PASSWORD
    if '$' in LOGIN_PASSWORD or LOGIN_PASSWORD == _DEFAULT_LOGIN_PASSWORD:
        return
    try:
        new_stored = _hash_password(plain_password)
        with open("password_override.txt", "w") as f:
            f.write(new_stored)
        LOGIN_PASSWORD = new_stored
    except Exception:
        pass


def _looks_like_valid_stored_password(stored):
    """
    Vérifie que le contenu de password_override.txt ressemble à un format stocké valide
    (haché "sel$empreinte", ou ancien format en clair) -- pour qu'un fichier corrompu (ex:
    écriture interrompue par une coupure serveur) fasse un repli sûr sur le mot de passe par
    défaut plutôt que de verrouiller l'accès administrateur sans aucun recours.
    """
    if '$' not in stored:
        return True  # ancien format en clair : tout contenu non vide est valide tel quel
    parts = stored.split('$', 1)
    if len(parts) != 2:
        return False
    salt, digest = parts
    try:
        bytes.fromhex(salt)
        bytes.fromhex(digest)
    except ValueError:
        return False
    return len(salt) == 32 and len(digest) == 64  # 16 octets de sel, empreinte SHA-256 (256 bits)


# LOGIN_PASSWORD contient la forme STOCKÉE (hachée si possible) du mot de passe courant --
# jamais le mot de passe en clair une fois qu'il a été changé au moins une fois.
# Recours en cas de mot de passe perdu ou de fichier corrompu : supprimer
# password_override.txt sur le serveur et redémarrer réinitialise au mot de passe par
# défaut d'usine (voir Notice, section Maintenance).
LOGIN_PASSWORD = _DEFAULT_LOGIN_PASSWORD
try:
    with open("password_override.txt", "r") as f:
        _saved_pw = f.read().strip()
        if _saved_pw and _looks_like_valid_stored_password(_saved_pw):
            LOGIN_PASSWORD = _saved_pw
        elif _saved_pw:
            print("⚠️  password_override.txt semble corrompu (format invalide) -- repli sur le mot de passe par défaut d'usine. Reconnectez-vous avec celui-ci puis changez-le à nouveau depuis l'interface.")
except FileNotFoundError:
    pass
SESSION_HOURS  = 8             # Durée de session en heures
MAX_LOGIN_ATTEMPTS = 5         # Tentatives max avant blocage 15 min
LOGIN_ATTEMPTS = {}            # {ip: [timestamps]}
_last_login_attempts_cleanup = time.time()


def _cleanup_login_attempts(now):
    """
    LOGIN_ATTEMPTS grossissait indéfiniment : chaque IP ayant un jour tenté de se connecter
    laissait une entrée à vie dans le dictionnaire (la route login() ne nettoyait que l'IP
    de la requête en cours, jamais les autres). Un balayage global, au plus une fois toutes
    les 5 minutes pour rester léger, retire les IP dont toutes les tentatives ont plus de
    15 minutes.
    """
    global _last_login_attempts_cleanup
    if now - _last_login_attempts_cleanup < 300:
        return
    _last_login_attempts_cleanup = now
    stale_ips = [ip for ip, timestamps in LOGIN_ATTEMPTS.items() if not any(now - t < 900 for t in timestamps)]
    for ip in stale_ips:
        LOGIN_ATTEMPTS.pop(ip, None)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not flask_session.get('logged_in'):
            # Routes API : renvoyer JSON 401 au lieu d'un redirect HTML
            if request.path.startswith('/api/'):
                return jsonify({"error": "Non authentifié", "redirect": "/login"}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def login_or_share_token_required(f):
    """
    Comme login_required, mais accepte aussi un lien de partage en lecture seule valide
    (paramètre ?token=... vérifié en base, avec expiration) -- utilisé uniquement pour les
    routes explicitement conçues pour être partagées (ex: page Synthèse de campagne).
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        if flask_session.get('logged_in'):
            return f(*args, **kwargs)
        token = request.args.get("token", "")
        if token and _is_valid_share_token(token):
            return f(*args, **kwargs)
        if request.path.startswith('/api/'):
            return jsonify({"error": "Non authentifié", "redirect": "/login"}), 401
        return redirect(url_for('login'))
    return decorated

# --- ENREGISTREMENT DU BLUEPRINT DES INTERVENTIONS ---
app.register_blueprint(interventions_bp)

# --- ENREGISTREMENT DU BLUEPRINT NDVI (extrait le 2026, cf. ndvi_bp.py pour le detail) ---
from ndvi_bp import ndvi_bp
app.register_blueprint(ndvi_bp)

# --- ENREGISTREMENT DU BLUEPRINT E-PHY (extrait le 2026, cf. ephy_bp.py pour le detail) ---
from ephy_bp import ephy_bp
app.register_blueprint(ephy_bp)

# --- ENREGISTREMENT DU BLUEPRINT MODELES D'INTERVENTION (cf. templates_bp.py) ---
from templates_bp import templates_bp
app.register_blueprint(templates_bp)

# --- ENREGISTREMENT DU BLUEPRINT LIENS DE PARTAGE (cf. share_tokens_bp.py) ---
# _is_valid_share_token est aussi importee ici pour le decorateur login_or_share_token_required
# (defini plus haut dans ce fichier), qui en a besoin pour verifier un token present dans l'URL.
from share_tokens_bp import share_tokens_bp, _is_valid_share_token
app.register_blueprint(share_tokens_bp)

# --- ENREGISTREMENT DU BLUEPRINT POINTS D'OBSERVATION TERRAIN (cf. field_points_bp.py) ---
from field_points_bp import field_points_bp
app.register_blueprint(field_points_bp)

# --- ENREGISTREMENT DU BLUEPRINT CAHIER DE TRACABILITE (cf. cahier_bp.py) ---
from cahier_bp import cahier_bp
app.register_blueprint(cahier_bp)

# --- ENREGISTREMENT DU BLUEPRINT EXPORTS RAPPORT DE CHANTIER (cf. chantier_export_bp.py) ---
from chantier_export_bp import chantier_export_bp
app.register_blueprint(chantier_export_bp)

# --- ENREGISTREMENT DU BLUEPRINT ANALYTIQUE (cf. analytique_bp.py) ---
from analytique_bp import analytique_bp
app.register_blueprint(analytique_bp)

# --- ENREGISTREMENT DU BLUEPRINT ASSISTANT D'IMPORT TRACCAR (cf. traccar_import_bp.py) ---
from traccar_import_bp import traccar_import_bp
app.register_blueprint(traccar_import_bp)

# --- ENREGISTREMENT DU BLUEPRINT SAUVEGARDE/RESTAURATION CONFIG (cf. config_backup_bp.py) ---
from config_backup_bp import config_backup_bp
app.register_blueprint(config_backup_bp)

# --- SAUVEGARDE AUTOMATIQUE DE LA BASE (cf. backup.py -- module utilitaire, pas de blueprint) ---
from backup import backup_database, backup_scheduler

# --- ENREGISTREMENT DU BLUEPRINT FERTILISATION (cf. fertilisation_bp.py) ---
from fertilisation_bp import fertilisation_bp
app.register_blueprint(fertilisation_bp)

from azote_bp import azote_bp
app.register_blueprint(azote_bp)

# ================= CONFIGURATION =================
# ── Configuration Traccar (chargée depuis config.json si disponible) ──
def _load_traccar_config():
    import json, os
    # Aucun identifiant par defaut : la configuration Traccar (URL, utilisateur, mot de
    # passe) doit etre explicitement renseignee dans config.json (ou via l'interface
    # "Parametrage Traccar"). Avant, des identifiants de production etaient codes en dur
    # ici comme valeurs par defaut -- retire pour eviter toute fuite si ce fichier est
    # partage ou versionne.
    defaults = {
        "url":      "",
        "user":     "",
        "password": "",
        "days_back": 30,
        "cache_duration": 60,
    }
    try:
        if os.path.exists("config.json"):
            with open("config.json", "r") as f:
                data = json.load(f)
                defaults.update(data.get("traccar", {}))
    except Exception:
        pass
    return defaults

_tcfg = _load_traccar_config()
TRACCAR_URL      = _tcfg["url"]
TRACCAR_USER     = _tcfg["user"]
TRACCAR_PASSWORD = _tcfg["password"]

DAYS_BACK      = _tcfg.get("days_back", 30)
CACHE_DURATION = _tcfg.get("cache_duration", 60)

if not TRACCAR_URL:
    import sys
    print(
        "⚠️  Aucune configuration Traccar trouvee (config.json manquant ou incomplet). "
        "Renseignez-la via l'icone 🛰️ Parametrage Traccar une fois connecte, ou en creant "
        "config.json avec une section \"traccar\": {\"url\":..., \"user\":..., \"password\":...}.",
        file=sys.stderr
    )

# ── Configuration des alertes (chargée depuis config.json si disponible) ──
def _load_alerts_config():
    defaults = {
        "max_working_speed_kmh": 15,   # au-delà : alerte vitesse anormale pendant le travail
        "min_coverage_pct": 85,        # en-deçà : alerte couverture insuffisante
        "jours_sans_intervention": 21, # au-delà : alerte parcelle sans intervention depuis longtemps
        "arret_prolonge_minutes": 5,   # au-delà : alerte arrêt prolongé outil engagé
    }
    try:
        if os.path.exists("config.json"):
            with open("config.json", "r") as f:
                data = json.load(f)
                defaults.update(data.get("alerts", {}))
    except Exception:
        pass
    return defaults

_alertscfg = _load_alerts_config()
ALERT_MAX_WORKING_SPEED_KMH = _alertscfg["max_working_speed_kmh"]
ALERT_MIN_COVERAGE_PCT      = _alertscfg["min_coverage_pct"]
ALERT_JOURS_SANS_INTERVENTION = _alertscfg["jours_sans_intervention"]
ALERT_ARRET_PROLONGE_MINUTES = _alertscfg["arret_prolonge_minutes"]

session = requests.Session()
session.auth = (TRACCAR_USER, TRACCAR_PASSWORD)
HEADERS = {"Accept": "application/json"}

_cached_data = None
_last_cache_time = 0

# ================= SAFE API GET =================
def safe_get(url):
    try:
        r = session.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 200:
            return r.json()
        return []
    except Exception:
        return []

# ================= CONVERSION BOOLEENNE ROBUSTE =================
def to_bool(val, default=False):
    """
    Convertit une valeur d'attribut Traccar en booleen, en gerant correctement :
    - un vrai booleen JSON (True/False)
    - une chaine de caracteres ("true"/"false"/"1"/"0", quelle que soit la casse)
    - un nombre (0/1)
    ATTENTION : ne JAMAIS faire bool(val) directement sur un attribut Traccar. Si l'appareil
    envoie isWorking sous forme de chaine (tres frequent selon le protocole/parseur), la chaine
    "false" est une chaine NON VIDE, donc bool("false") vaut True en Python -- un piege classique
    qui inversait silencieusement l'etat de travail (tout etait considere "actif").
    """
    if isinstance(val, bool):
        return val
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    if s in ("true", "1", "yes", "on", "oui"):
        return True
    if s in ("false", "0", "no", "off", "non", ""):
        return False
    return default

# ================= FORMATAGE DATE =================
def format_date_fr(dt_str):
    if not dt_str: return ""
    try:
        dt = datetime.strptime(dt_str[:19], "%Y-%m-%dT%H:%M:%S")
        return dt.strftime("%d/%m/%Y %Hh%M")
    except Exception:
        return dt_str

# ================= CORE EXTRACTION =================
def get_devices():
    data = safe_get(f"{TRACCAR_URL}/devices")
    return {d["id"]: d for d in data if "id" in d} if isinstance(data, list) else {}

def get_geofences():
    data = safe_get(f"{TRACCAR_URL}/geofences")
    return {g["id"]: g for g in data if "id" in g} if isinstance(data, list) else {}

def fetch_positions_for_device(args):
    d_id, start_str, end_str = args
    url = f"{TRACCAR_URL}/reports/route?deviceId={d_id}&from={start_str}&to={end_str}"
    data = safe_get(url)
    if isinstance(data, list):
        for p in data:
            p["deviceId"] = d_id
        return data
    return []

def fetch_events_for_device(args):
    d_id, start_str, end_str = args
    url = f"{TRACCAR_URL}/reports/events?deviceId={d_id}&from={start_str}&to={end_str}"
    data = safe_get(url)
    if isinstance(data, list):
        return [e for e in data if "geo" in str(e.get("type", "")).lower() or e.get("geofenceId")]
    return []

def get_data_parallel(device_ids):
    end = datetime.utcnow()
    start = end - timedelta(days=DAYS_BACK)
    start_str = start.strftime('%Y-%m-%dT%H:%M:%SZ')
    end_str = end.strftime('%Y-%m-%dT%H:%M:%SZ')
    
    tasks = [(d_id, start_str, end_str) for d_id in device_ids]
    all_positions = []
    all_events = []
    
    with ThreadPoolExecutor(max_workers=10) as executor:
        pos_results = executor.map(fetch_positions_for_device, tasks)
        evt_results = executor.map(fetch_events_for_device, tasks)
        
        for res in pos_results: all_positions.extend(res)
        for res in evt_results: all_events.extend(res)
            
    return sorted(all_positions, key=lambda x: x.get("fixTime", "")), all_events

def build_positions_index(positions):
    index = {}
    for p in positions:
        d_id = p.get("deviceId")
        if d_id: index.setdefault(d_id, []).append(p)
    return index

def find_position(device_id, event_time, positions_index):
    plist = positions_index.get(device_id, [])
    if not plist: return {}
    try:
        et = datetime.strptime(event_time[:19], "%Y-%m-%dT%H:%M:%S")
    except Exception: return {}
    best, best_diff = {}, float("inf")
    for p in plist:
        try:
            pt = datetime.strptime(p.get("fixTime","")[:19], "%Y-%m-%dT%H:%M:%S")
            diff = abs((pt - et).total_seconds())
            if diff < best_diff:
                best, best_diff = p, diff
        except Exception: continue
    return best if best_diff <= 600 else {}

def build_last_positions(positions):
    last = {}
    for p in positions:
        d = p.get("deviceId")
        if not d: continue
        if d not in last or p.get("fixTime", "") > last[d].get("fixTime", ""):
            last[d] = p
    return last

# ================= CACHE CONTROL & CALCUL DURÉES =================
def build_data():
    global _cached_data, _last_cache_time
    current_time = time.time()

    if _cached_data and (current_time - _last_cache_time < CACHE_DURATION):
        return _cached_data

    devices = get_devices()
    geofences = get_geofences()
    
    if not devices:
        return {"events": [], "positions": [], "geofences": {}}

    positions, events = get_data_parallel(list(devices.keys()))
    positions_index = build_positions_index(positions)
    last_positions = build_last_positions(positions)

    raw_events = []
    for e in events:
        device_id = e.get("deviceId")
        geo_id = e.get("geofenceId")
        pos = find_position(device_id, e.get("eventTime"), positions_index)
        attrs = pos.get("attributes") or {}
        etype = "Entrée" if "enter" in str(e.get("type","")).lower() else "Sortie"

        w_val = str(attrs.get("workingWidth", attrs.get("width", ""))).strip()
        width_str = ""
        if w_val and w_val != "None":
            if "m" in w_val.lower():
                width_str = w_val
            else:
                width_str = f"{w_val} m"

        # Capturer appliedArea de manière indépendante de field
        applied_area_val = attrs.get("appliedArea", "")

        raw_events.append({
            "deviceId": device_id,
            "geofenceId": geo_id,
            "vehicle": devices.get(device_id, {}).get("name", f"Véhicule {device_id}"),
            "geofence": geofences.get(geo_id, {}).get("name", f"Parcelle {geo_id}"),
            "type": etype,
            "date": e.get("eventTime"),
            "date_fr": format_date_fr(e.get("eventTime")),
            "field": attrs.get("field", ""),
            "appliedArea": str(applied_area_val).strip() if applied_area_val is not None else "",
            "tool": attrs.get("tool", ""),
            "width": width_str,
            "lat": pos.get("latitude", ""),
            "lon": pos.get("longitude", ""),
        })

    raw_events.sort(key=lambda x: x["date"])
    active_inputs = {} 
    
    for e in raw_events:
        key = (e["deviceId"], e["geofenceId"])
        if e["type"] == "Entrée":
            active_inputs[key] = e["date"]
            e["duration"] = "-"
        elif e["type"] == "Sortie":
            if key in active_inputs:
                try:
                    t_in = datetime.strptime(active_inputs[key][:19], "%Y-%m-%dT%H:%M:%S")
                    t_out = datetime.strptime(e["date"][:19], "%Y-%m-%dT%H:%M:%S")
                    diff = t_out - t_in
                    hours, remainder = divmod(int(diff.total_seconds()), 3600)
                    minutes, _ = divmod(remainder, 60)
                    e["duration"] = f"{hours}h{minutes:02d}m" if hours > 0 else f"{minutes}m"
                    del active_inputs[key]
                except Exception:
                    e["duration"] = "-"
            else:
                e["duration"] = "-"

    _cached_data = {
        "events": raw_events,
        "positions": [
            {
                "vehicle": devices.get(p.get("deviceId"), {}).get("name", "Inconnu"),
                "lat": p.get("latitude"),
                "lon": p.get("longitude"),
                "date": p.get("fixTime")
            }
            for p in last_positions.values()
        ],
        "geofences": geofences
    }
    _last_cache_time = current_time
    print(f"✅ Cache v7.14 mis à jour (Surfaces distinctes : field & appliedArea).")
    return _cached_data

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    ip = request.remote_addr
    now = time.time()

    _cleanup_login_attempts(now)

    # Nettoyer les anciennes tentatives (> 15 min) pour cette IP
    LOGIN_ATTEMPTS[ip] = [t for t in LOGIN_ATTEMPTS.get(ip, []) if now - t < 900]
    if not LOGIN_ATTEMPTS[ip]:
        LOGIN_ATTEMPTS.pop(ip, None)

    if len(LOGIN_ATTEMPTS.get(ip, [])) >= MAX_LOGIN_ATTEMPTS:
        wait_min = int((900 - (now - LOGIN_ATTEMPTS[ip][0])) / 60) + 1
        error = f"Trop de tentatives. Réessayez dans {wait_min} min."
        return render_template("login.html", error=error)

    if request.method == "POST":
        if request.form.get("username") == LOGIN_USER and _verify_password(request.form.get("password", ""), LOGIN_PASSWORD):
            LOGIN_ATTEMPTS.pop(ip, None)
            _migrate_legacy_plaintext_password(request.form.get("password", ""))
            flask_session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=SESSION_HOURS)
            flask_session["logged_in"] = True
            return redirect(url_for("index"))
        LOGIN_ATTEMPTS.setdefault(ip, []).append(now)
        remaining = MAX_LOGIN_ATTEMPTS - len(LOGIN_ATTEMPTS[ip])
        error = f"Identifiants incorrects. ({remaining} tentative(s) restante(s))" if remaining > 0 else "Trop de tentatives. Réessayez dans 15 min."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    flask_session.clear()
    return redirect(url_for("login"))

@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    global LOGIN_PASSWORD
    error = None
    success = None
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not _verify_password(current, LOGIN_PASSWORD):
            error = "Mot de passe actuel incorrect."
        elif len(new_pw) < 6:
            error = "Le nouveau mot de passe doit faire au moins 6 caractères."
        elif new_pw != confirm:
            error = "Les deux mots de passe ne correspondent pas."
        else:
            new_stored = _hash_password(new_pw)
            LOGIN_PASSWORD = new_stored
            # Persister (haché, jamais en clair) dans un fichier pour survivre au redémarrage
            try:
                with open("password_override.txt", "w") as f:
                    f.write(new_stored)
            except Exception:
                pass
            success = "Mot de passe mis à jour avec succès."
    return render_template("change_password.html", error=error, success=success)

@app.route("/")
@login_required
def index():
    return render_template("index.html")



# Assistant d'import Traccar (page, config, proxy, historique) : extrait dans
# traccar_import_bp.py, voir l'enregistrement du blueprint en tete de fichier.



# ================= SOUS-ZONES (scission d'une parcelle en plusieurs cultures) =================
# Une parcelle Traccar (geofence_id) reste l'unique référence pour tout l'historique existant
# (interventions, carte de chantier, GPS...) : on ne redessine JAMAIS la géofence elle-même.
# Une "sous-parcelle" est une donnée propre à l'application, superposée à une parcelle pour une
# campagne donnée, quand celle-ci est scindée en plusieurs cultures la même année. Sans
# sous-parcelle définie, tout se comporte exactement comme avant (comportement par défaut inchangé).
def _ensure_sous_parcelles_table():
    with sqlite3.connect('database.db') as conn:
        # Migration de compatibilité : la fonctionnalité s'appelait "sous-zone" à sa création
        # et a été renommée en "sous-parcelle" -- si l'ancienne table/colonne existe encore
        # (installation mise à jour depuis une version antérieure), on les renomme plutôt que
        # d'en recréer de nouvelles vides, pour ne perdre aucune donnée déjà saisie.
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        if 'sous_zones' in tables and 'sous_parcelles' not in tables:
            conn.execute("ALTER TABLE sous_zones RENAME TO sous_parcelles")
        cols_interv = {row[1] for row in conn.execute("PRAGMA table_info(interventions)").fetchall()}
        if 'sous_zone_id' in cols_interv and 'sous_parcelle_id' not in cols_interv:
            conn.execute("ALTER TABLE interventions RENAME COLUMN sous_zone_id TO sous_parcelle_id")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS sous_parcelles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                geofence_id TEXT NOT NULL,
                nom TEXT NOT NULL,
                campagne TEXT,
                culture TEXT,
                statut TEXT DEFAULT 'attente',
                surface_ha REAL,
                polygon TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        # Colonne de rattachement optionnel sur les interventions (NULL = pas de sous-parcelle,
        # comportement historique inchangé).
        cols = {row[1] for row in conn.execute("PRAGMA table_info(interventions)").fetchall()}
        if 'sous_parcelle_id' not in cols:
            conn.execute("ALTER TABLE interventions ADD COLUMN sous_parcelle_id INTEGER DEFAULT NULL")


def _get_sous_parcelles_info(conn, geofence_id=None):
    """
    Renvoie {sous_parcelle_id: {id, geofence_id, nom, culture, campagne, surface_ha,
    polygon}} pour toutes les sous-parcelles (ou seulement celles d'une géofence donnée si
    geofence_id est précisé). Centralise cette requête : elle était dupliquée par le passé
    dans 7 endroits différents du code (Bilan, Fertilisation, export PDF, Cahier de
    traçabilité, anomalies, phénologie maïs/céréales, Analytique), chacun avec un jeu de
    colonnes parfois incomplet selon l'endroit -- un vrai risque d'incohérence à chaque
    nouvelle fonctionnalité qui y touchait sans se souvenir de reproduire exactement la
    bonne requête partout. N'appelle PAS _ensure_sous_parcelles_table() elle-même : à faire
    par l'appelant, qui sait s'il vient de créer la connexion ou pas.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    if geofence_id is not None:
        cur.execute(
            "SELECT id, geofence_id, nom, culture, campagne, surface_ha, polygon FROM sous_parcelles WHERE geofence_id = ?",
            (str(geofence_id),)
        )
    else:
        cur.execute("SELECT id, geofence_id, nom, culture, campagne, surface_ha, polygon FROM sous_parcelles")
    return {r["id"]: dict(r) for r in cur.fetchall()}


# api_sous_parcelles / api_sous_parcelle_detail : code mort retire (confirme par test
# empirique) -- la version reellement active est celle du blueprint interventions_bp,
# enregistree plus tot et donc prioritaire pour ces memes routes.

# Points d'observation terrain : extrait dans field_points_bp.py, voir l'enregistrement
# du blueprint en tete de fichier.



# ================= MODELES D'INTERVENTION (produits/doses réutilisables) =================
# ================= LIENS DE PARTAGE EN LECTURE SEULE =================
# Liens de partage en lecture seule : extrait dans share_tokens_bp.py (voir l'import et
# l'enregistrement du blueprint en tete de fichier).



# Modeles d'intervention reutilisables : extrait dans templates_bp.py, voir
# l'enregistrement du blueprint en tete de fichier.



@app.route("/synthese")
@login_or_share_token_required
def synthese():
    """Page dédiée : tableau de bord synthétique (météo + avancement par parcelle)."""
    return render_template("synthese.html")


@app.route("/api/synthese_campagne")
@login_or_share_token_required
def api_synthese_campagne():
    """
    Agrège des données déjà calculées ailleurs (statuts parcelles, dernière intervention,
    alertes DAR, alertes opérationnelles du jour, position GPS pour la météo) pour la page
    /synthese. Ne recalcule PAS de couverture GPS cumulée sur toute la campagne (trop coûteux
    à l'échelle de toutes les parcelles) : l'avancement affiché se base sur le statut de chaque
    parcelle (attente/préparé/semé/traité/récolté), déjà tenu à jour par le reste de
    l'application. Les alertes vitesse/arrêt prolongé, elles, sont recalculées mais bornées à
    la journée en cours (pas de cumul historique), ce qui reste rapide même avec beaucoup de
    véhicules.
    """
    return jsonify(_compute_synthese_campagne())


def _compute_synthese_campagne():
    """Calcul partagé entre /api/synthese_campagne (JSON) et l'export PDF de synthèse."""
    DB_PATH = 'database.db'
    now = datetime.now()
    campagne_actuelle = str(now.year)
    _ensure_sous_parcelles_table()

    with sqlite3.connect(DB_PATH) as conn:
        # Recalcule les statuts automatiques (parcelles ET sous-parcelles) avant de les lire,
        # pour que la Synthèse soit toujours à jour même si un chemin de sauvegarde
        # d'intervention n'aurait pas explicitement déclenché ce recalcul par ailleurs.
        _auto_update_statuts(conn)

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT geofence_id, identifiant, nom_parcelle, statut, surface_ha FROM parcelles")
        parcelles = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT geofence_id, MAX(exit_time) AS derniere, COUNT(*) AS nb FROM interventions GROUP BY geofence_id")
        dern = {row["geofence_id"]: {"derniere": row["derniere"], "nb": row["nb"]} for row in cur.fetchall()}

        # Sous-parcelles actives pour la campagne en cours : quand une parcelle a été scindée
        # en plusieurs cultures, elle est représentée par une carte par sous-parcelle plutôt
        # qu'une seule carte ambiguë pour toute la parcelle. La campagne est toujours désignée
        # par son année de RÉCOLTE (ex: un blé semé en 2025 et récolté en 2026 appartient à la
        # "campagne 2026", jamais "2025-2026"). Le filtre en LIKE reste néanmoins tolérant, en
        # repli, si le champ "campagne" d'une sous-parcelle a été saisi à la main sous un
        # format différent (texte libre) plutôt que via l'année de récolte attendue.
        cur.execute("SELECT * FROM sous_parcelles WHERE campagne LIKE ?", (f"%{campagne_actuelle}%",))
        sous_parcelles_par_geofence = {}
        for r in cur.fetchall():
            sous_parcelles_par_geofence.setdefault(str(r["geofence_id"]), []).append(dict(r))

        # Stats (dernière intervention, nombre) spécifiques à chaque sous-parcelle, quand des
        # interventions y ont été explicitement rattachées (sous_parcelle_id renseigné).
        cur.execute("""
            SELECT geofence_id, sous_parcelle_id, MAX(exit_time) AS derniere, COUNT(*) AS nb
            FROM interventions WHERE sous_parcelle_id IS NOT NULL GROUP BY geofence_id, sous_parcelle_id
        """)
        dern_sous_parcelle = {(str(row["geofence_id"]), row["sous_parcelle_id"]): {"derniere": row["derniere"], "nb": row["nb"]} for row in cur.fetchall()}

        # Véhicule + date de la toute dernière intervention par parcelle, pour permettre un
        # lien direct vers la carte de chantier sur le bon jour et le bon véhicule (plutôt
        # que d'ouvrir la carte "à vide" en laissant l'utilisateur tout resélectionner).
        cur.execute("""
            SELECT i1.geofence_id, i1.device_id, i1.exit_time
            FROM interventions i1
            INNER JOIN (
                SELECT geofence_id, MAX(exit_time) AS max_exit FROM interventions GROUP BY geofence_id
            ) i2 ON i1.geofence_id = i2.geofence_id AND i1.exit_time = i2.max_exit
        """)
        dernier_detail = {row["geofence_id"]: {"device_id": row["device_id"], "exit_time": row["exit_time"]} for row in cur.fetchall()}

    raw = build_data()
    geofences = raw.get("geofences", {})

    parcelles_eclatees = []
    for p in parcelles:
        p["nom_traccar"] = resolve_geo_name(geofences, p["geofence_id"])
        geo_id_str = str(p["geofence_id"])
        sous_parcelles = sous_parcelles_par_geofence.get(geo_id_str)

        if sous_parcelles:
            # Cette parcelle est scindée pour la campagne en cours : une carte par sous-parcelle,
            # chacune avec son propre statut/culture/surface, à la place de la carte unique.
            for sz in sous_parcelles:
                info_sz = dern_sous_parcelle.get((geo_id_str, sz["id"]))
                jours_sz = None
                if info_sz and info_sz.get("derniere"):
                    try:
                        jours_sz = (now - datetime.strptime(info_sz["derniere"][:19], "%Y-%m-%dT%H:%M:%S")).days
                    except Exception:
                        jours_sz = None
                parcelles_eclatees.append({
                    "geofence_id": p["geofence_id"],
                    "sous_parcelle_id": sz["id"],
                    "identifiant": p.get("identifiant"),
                    "nom_traccar": f"{p['nom_traccar']} — {sz['nom']}",
                    "nom_parcelle": f"{p.get('nom_parcelle') or p['nom_traccar']} — {sz['nom']}",
                    "statut": sz.get("statut"),
                    "culture": sz.get("culture"),
                    "surface_ha": sz.get("surface_ha"),
                    "jours_sans_intervention": jours_sz,
                    "nb_interventions": info_sz["nb"] if info_sz else 0,
                    "derniere_date": None, "derniere_device_id": None,
                    "est_sous_parcelle": True,
                })
            continue

        info = dern.get(p["geofence_id"])
        jours = None
        if info and info.get("derniere"):
            try:
                jours = (now - datetime.strptime(info["derniere"][:19], "%Y-%m-%dT%H:%M:%S")).days
            except Exception:
                jours = None
        p["jours_sans_intervention"] = jours
        p["nb_interventions"] = info["nb"] if info else 0

        detail = dernier_detail.get(p["geofence_id"])
        if detail and detail.get("device_id") not in (None, 0, "0"):
            p["derniere_date"] = str(detail["exit_time"])[:10]
            p["derniere_device_id"] = detail["device_id"]
        else:
            p["derniere_date"] = None
            p["derniere_device_id"] = None

    # Fusionne : les parcelles scindées sont remplacées par leurs sous-parcelles (déjà ajoutées
    # à parcelles_eclatees plus haut, via "continue"), les autres restent inchangées.
    parcelles = parcelles_eclatees + [p for p in parcelles if str(p["geofence_id"]) not in sous_parcelles_par_geofence]

    statut_counts = {}
    for p in parcelles:
        s = p.get("statut") or ""
        statut_counts[s] = statut_counts.get(s, 0) + 1

    try:
        alertes_data = alertes_dar().get_json()
    except Exception:
        alertes_data = {"alertes": [], "nb_alertes": 0}

    # Alertes opérationnelles du jour (vitesse anormale + arrêts prolongés), déjà calculées
    # par le même moteur que la carte de chantier -- bornées à la journée en cours (pas de
    # cumul sur toute la campagne, pour rester rapide même avec beaucoup de véhicules).
    import math
    today_str = now.strftime("%Y-%m-%d")
    devices = get_devices()
    device_ids = list(devices.keys())[:40]  # garde-fou : évite un nombre de véhicules démesuré

    # Géométries des parcelles déjà connues côté Traccar, pour attribuer chaque point
    # d'alerte à la bonne parcelle (affichage d'un badge directement sur sa carte Kanban).
    raw_geofences_list = safe_get(f"{TRACCAR_URL}/geofences")
    parsed_geofences = {}
    if isinstance(raw_geofences_list, list):
        for g in raw_geofences_list:
            geom = _parse_geofence_wkt(g.get("area"))
            if geom:
                parsed_geofences[g["id"]] = geom

    def _geofence_id_for_point(lat, lon):
        for gid, geom in parsed_geofences.items():
            if geom["type"] == "polygon" and _point_in_polygon(lat, lon, geom["coords"]):
                return gid
            if geom["type"] == "circle" and _haversine_m(lat, lon, geom["center"][0], geom["center"][1]) <= geom["radius"]:
                return gid
        return None

    nb_alertes_vitesse = 0
    nb_alertes_arret = 0
    geofences_avec_alerte_operationnelle = set()

    def _device_alert_counts(dev_id):
        try:
            pts = _fetch_and_parse_positions(dev_id, today_str)
            if not pts:
                return (0, 0, set())
            data = _build_corridors_response(pts, today_str, math)
            alerts = data.get("alerts", {})
            vit = alerts.get("vitesse", [])
            arr = alerts.get("arrets_prolonges", [])
            touched = set()
            for pt in (vit + arr):
                gid = _geofence_id_for_point(pt["lat"], pt["lon"])
                if gid is not None:
                    touched.add(gid)
            return (len(vit), len(arr), touched)
        except Exception:
            return (0, 0, set())

    if device_ids:
        with ThreadPoolExecutor(max_workers=8) as executor:
            for v, a, touched in executor.map(_device_alert_counts, device_ids):
                nb_alertes_vitesse += v
                nb_alertes_arret += a
                geofences_avec_alerte_operationnelle |= touched

    # Géofences concernées par une alerte DAR active (déjà calculée plus haut)
    geofences_avec_alerte_dar = {a["geofence_id"] for a in alertes_data.get("alertes", [])}
    geofences_avec_alerte = {str(g) for g in (geofences_avec_alerte_dar | geofences_avec_alerte_operationnelle)}
    for p in parcelles:
        p["a_une_alerte"] = str(p["geofence_id"]) in geofences_avec_alerte

    first_position = None
    for pos in raw.get("positions", []):
        if pos.get("lat") and pos.get("lon"):
            first_position = {"lat": pos["lat"], "lon": pos["lon"]}
            break

    # Repli : si build_data() n'a trouvé aucune position récente (dépend de l'historique
    # des 30 derniers jours), on interroge directement l'endpoint Traccar des positions
    # "live" -- plus fiable pour simplement obtenir un point de repère pour la météo.
    if not first_position:
        try:
            live_positions = safe_get(f"{TRACCAR_URL}/positions")
            if isinstance(live_positions, list):
                for p in live_positions:
                    if p.get("latitude") and p.get("longitude"):
                        first_position = {"lat": p["latitude"], "lon": p["longitude"]}
                        break
        except Exception:
            pass

    nb_alertes_total = alertes_data.get("nb_alertes", 0) + nb_alertes_vitesse + nb_alertes_arret

    return {
        "date": now.strftime("%Y-%m-%d"),
        "parcelles": parcelles,
        "statut_counts": statut_counts,
        "nb_alertes_dar": alertes_data.get("nb_alertes", 0),
        "alertes_dar": alertes_data.get("alertes", []),
        "nb_alertes_vitesse": nb_alertes_vitesse,
        "nb_alertes_arret": nb_alertes_arret,
        "nb_alertes_total": nb_alertes_total,
        "first_position": first_position,
    }


@app.route("/aide")
@login_required
def aide():
    import os
    for p in [os.path.join(app.template_folder or "templates", "Notice.html"),
              os.path.join(os.path.dirname(__file__), "Notice.html"),
              "Notice.html"]:
        if os.path.isfile(p):
            return send_file(p)
    return render_template("Notice.html")

def apply_filters(events, vehicle, geofence, start, end):
    def parse_dt(s):
        if not s: return None
        try: return datetime.strptime(s[:16], "%Y-%m-%dT%H:%M")
        except Exception:
            try: return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
            except Exception: return None

    filtered = events
    if start:
        s = parse_dt(start)
        if s: filtered = [e for e in filtered if parse_dt(e.get("date")) and parse_dt(e.get("date")) >= s]
    if end:
        e_dt = parse_dt(end)
        if e_dt: filtered = [e for e in filtered if parse_dt(e.get("date")) and parse_dt(e.get("date")) <= e_dt]
    if vehicle:
        filtered = [e for e in filtered if e["vehicle"] == vehicle]
    if geofence:
        filtered = [e for e in filtered if e["geofence"] == geofence]
    return filtered

STATUT_PAR_TYPE_INTERVENTION = {
    'Semis': 'semis',
    'Pulvérisation': 'traite',
    'Épandage': 'traite',
    'Récolte': 'recolte',
    'Labour': 'prepare',
    'Hersage': 'prepare',
    'Déchaumage': 'prepare',
    'Broyage': 'prepare',
    'Travail du sol': 'prepare',
}


def _auto_update_statuts(conn):
    """
    Recalcule le statut automatique de chaque parcelle (mode auto uniquement) et de chaque
    sous-parcelle, à partir de leur dernière intervention rattachée. Fonction réutilisable :
    appelée à la fois par la route /api/parcelles/refresh_statuts ET directement par la
    Synthèse de campagne, pour que l'affichage soit toujours à jour même si un chemin de
    sauvegarde d'intervention aurait oublié de déclencher le recalcul explicitement.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT geofence_id, intervention_type, exit_time
        FROM interventions
        ORDER BY exit_time ASC
    """)
    derniere_par_parcelle = {}
    for row in cur.fetchall():
        derniere_par_parcelle[row['geofence_id']] = row['intervention_type']

    cur.execute("SELECT geofence_id FROM parcelles WHERE statut_auto = 1 OR statut_auto IS NULL")
    parcelles_auto = [row[0] for row in cur.fetchall()]

    updated = 0
    for geo_id in parcelles_auto:
        type_interv = derniere_par_parcelle.get(geo_id)
        nouveau_statut = STATUT_PAR_TYPE_INTERVENTION.get(type_interv, 'attente')
        cur.execute(
            "UPDATE parcelles SET statut = ? WHERE geofence_id = ? AND statut_auto = 1",
            (nouveau_statut, geo_id)
        )
        updated += cur.rowcount

    cur.execute("""
        SELECT sous_parcelle_id, intervention_type, exit_time
        FROM interventions WHERE sous_parcelle_id IS NOT NULL
        ORDER BY exit_time ASC
    """)
    derniere_par_sous_parcelle = {}
    for row in cur.fetchall():
        derniere_par_sous_parcelle[row['sous_parcelle_id']] = row['intervention_type']

    for sp_id, type_interv in derniere_par_sous_parcelle.items():
        nouveau_statut = STATUT_PAR_TYPE_INTERVENTION.get(type_interv, 'attente')
        cur.execute("UPDATE sous_parcelles SET statut = ? WHERE id = ?", (nouveau_statut, sp_id))
        updated += cur.rowcount

    conn.commit()
    return updated


@app.route("/api/parcelles/refresh_statuts", methods=["POST"])
@login_required
def refresh_statuts():
    """
    Recalcule le statut automatique de chaque parcelle à partir de sa
    dernière intervention enregistrée. Ne touche pas aux parcelles en
    mode manuel (statut_auto = 0). Fait de même pour les sous-parcelles,
    à partir des interventions qui leur sont explicitement rattachées
    (sous_parcelle_id) -- sans quoi une sous-parcelle resterait figée
    indéfiniment au statut choisi à sa création.
    """
    DB_PATH = 'database.db'
    _ensure_sous_parcelles_table()
    with sqlite3.connect(DB_PATH) as conn:
        updated = _auto_update_statuts(conn)

    return jsonify({"status": "success", "updated": updated})


@app.route("/api/parcelles/nouvelle_campagne", methods=["POST"])
@login_required
def nouvelle_campagne():
    """
    Démarre une nouvelle campagne : remet le statut de TOUTES les parcelles à 'attente'
    (y compris celles en mode manuel, contrairement à refresh_statuts), pour repartir sur
    une saison vierge. Ne touche ni aux identifiants/surfaces des parcelles, ni à
    l'historique des interventions passées (toujours consultable dans le Carnet).
    """
    DB_PATH = 'database.db'
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.execute("UPDATE parcelles SET statut = 'attente'")
        updated = cur.rowcount
        conn.commit()
    return jsonify({"status": "success", "updated": updated})


def resolve_geo_name(geofences, geo_id):
    """geofences peut être indexé par int ou str selon la source : on tente les deux clés."""
    info = geofences.get(geo_id)
    if info is None:
        info = geofences.get(str(geo_id))
    if info is None:
        try:
            info = geofences.get(int(geo_id))
        except (ValueError, TypeError):
            info = None
    if isinstance(info, dict) and info.get('name'):
        return info['name']
    return f"Parcelle {geo_id}"


STATUT_INFO_PDF = {
    "":        {"label": "Non renseigne",  "color": "#94a3b8", "pct": 0},
    "attente": {"label": "En attente",     "color": "#d97706", "pct": 10},
    "prepare": {"label": "Prepare",        "color": "#7c3aed", "pct": 30},
    "semis":   {"label": "Seme",           "color": "#16a34a", "pct": 55},
    "traite":  {"label": "Traite",         "color": "#2563eb", "pct": 80},
    "recolte": {"label": "Recolte",        "color": "#dc2626", "pct": 100},
}

@app.route("/api/export_synthese_pdf")
@login_or_share_token_required
def api_export_synthese_pdf():
    """Genere un PDF de la synthese de campagne : avancement par parcelle, alertes, anomalies, phenologie."""
    data = _compute_synthese_campagne()
    try:
        anomalies = api_incoherences().get_json().get("anomalies", [])
    except Exception:
        anomalies = []

    try:
        pheno_mais = _compute_phenologie_mais().get("parcelles", [])
    except Exception:
        pheno_mais = []
    try:
        pheno_cereales = _compute_phenologie_cereales().get("parcelles", [])
    except Exception:
        pheno_cereales = []

    try:
        pdf_path = _generate_synthese_pdf(data, anomalies, pheno_mais, pheno_cereales)
    except Exception as e:
        app.logger.exception("export_synthese_pdf: erreur generation PDF")
        return jsonify({"error": f"Erreur generation PDF : {e}"}), 500

    filename = f"synthese_campagne_{data['date']}.pdf"
    return send_file(pdf_path, mimetype="application/pdf", as_attachment=True, download_name=filename)


def _generate_synthese_pdf(data, anomalies, pheno_mais=None, pheno_cereales=None):
    def safe(t):
        return str(t if t is not None else '').encode('latin-1', 'replace').decode('latin-1')

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    page_w = 190

    # En-tete
    date_fr = datetime.strptime(data["date"], "%Y-%m-%d").strftime("%d-%m-%Y")
    pdf.set_fill_color(14, 116, 144)  # cyan (#0e7490), couleur de la page Synthese
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 12, safe(f"Synthese de campagne - {date_fr}"), ln=1, fill=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # Stats globales
    nb_parcelles = len(data["parcelles"])
    avg_pct = round(sum(STATUT_INFO_PDF.get(p.get("statut") or "", STATUT_INFO_PDF[""])["pct"] for p in data["parcelles"]) / nb_parcelles) if nb_parcelles else 0

    pdf.set_font("Arial", "B", 11)
    pdf.set_fill_color(30, 41, 59); pdf.set_text_color(255, 255, 255)
    pdf.cell(page_w, 8, safe("  Vue d'ensemble"), ln=1, fill=True)
    pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 10)
    rows = [
        ("Parcelles suivies", str(nb_parcelles)),
        ("Avancement moyen de campagne", f"{avg_pct} %"),
        ("Alertes actives (DAR + vitesse + arret)", str(data.get("nb_alertes_total", 0))),
        ("dont alertes DAR", str(data.get("nb_alertes_dar", 0))),
        ("dont vitesse anormale (aujourd'hui)", str(data.get("nb_alertes_vitesse", 0))),
        ("dont arrets prolonges (aujourd'hui)", str(data.get("nb_alertes_arret", 0))),
    ]
    for label, val in rows:
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(page_w*0.65, 7, safe(f"  {label}"), border=1)
        pdf.cell(page_w*0.35, 7, safe(val), border=1, align="R")
        pdf.ln()
    pdf.ln(4)

    # Tableau par parcelle
    pdf.set_font("Arial", "B", 11)
    pdf.set_fill_color(30, 41, 59); pdf.set_text_color(255, 255, 255)
    pdf.cell(page_w, 8, safe("  Avancement par parcelle"), ln=1, fill=True)
    pdf.set_text_color(0, 0, 0)

    parcelles_sorted = sorted(
        data["parcelles"],
        key=lambda p: (p.get("nom_traccar") or p.get("nom_parcelle") or "").lower()
    )

    pdf.set_font("Arial", "B", 9)
    headers = ["Parcelle", "Statut", "Avanc.", "Surface", "Dernier travail"]
    widths  = [page_w*0.32, page_w*0.20, page_w*0.12, page_w*0.14, page_w*0.22]
    pdf.set_fill_color(226, 232, 240)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 7, safe(h), border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Arial", "", 8)
    for p in parcelles_sorted:
        info = STATUT_INFO_PDF.get(p.get("statut") or "", STATUT_INFO_PDF[""])
        nom = p.get("nom_traccar") or p.get("nom_parcelle") or f"Parcelle {p['geofence_id']}"
        alerte_marker = "! " if p.get("a_une_alerte") else ""
        jours = p.get("jours_sans_intervention")
        dernier = "jamais travaillee" if jours is None else f"{jours} j"
        r, g, b = _hex_to_rgb(info["color"])
        pdf.set_text_color(0, 0, 0)
        pdf.cell(widths[0], 6, safe(alerte_marker + nom), border=1)
        pdf.set_fill_color(r, g, b); pdf.set_text_color(255, 255, 255)
        pdf.cell(widths[1], 6, safe(info["label"]), border=1, align="C", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(widths[2], 6, safe(f"{info['pct']}%"), border=1, align="C")
        pdf.cell(widths[3], 6, safe(f"{p['surface_ha']} ha" if p.get("surface_ha") else "-"), border=1, align="C")
        pdf.cell(widths[4], 6, safe(dernier), border=1, align="C")
        pdf.ln()

    # Phénologie (maïs + céréales à paille), sous forme de tableau compact : stade
    # actuellement atteint + prochain stade avec sa date (réelle ou estimée).
    def _add_phenologie_table(title, parcelles):
        if not parcelles:
            return
        pdf.ln(4)
        pdf.set_font("Arial", "B", 11)
        pdf.set_fill_color(22, 101, 52); pdf.set_text_color(255, 255, 255)
        pdf.cell(page_w, 8, safe(f"  {title}"), ln=1, fill=True)
        pdf.set_text_color(0, 0, 0)

        pdf.set_font("Arial", "B", 8)
        headers2 = ["Parcelle", "Semence", "Semis", "Cumul DJ", "Stade actuel", "Prochain stade"]
        widths2 = [page_w*0.19, page_w*0.19, page_w*0.12, page_w*0.12, page_w*0.19, page_w*0.19]
        pdf.set_fill_color(226, 232, 240)
        for i, h in enumerate(headers2):
            pdf.cell(widths2[i], 6, safe(h), border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for p in parcelles:
            stages = p.get("stages", [])
            stage_label = next((s["label"] for s in stages if s["key"] == p.get("stage_atteint")), p.get("stage_atteint", "-"))
            idx = next((i for i, s in enumerate(stages) if s["key"] == p.get("stage_atteint")), None)
            nxt = stages[idx+1] if (idx is not None and idx+1 < len(stages)) else None
            if nxt and nxt.get("date"):
                date_fr = datetime.strptime(nxt["date"], "%Y-%m-%d").strftime("%d/%m/%Y")
                prefix = "~" if nxt.get("date_estimee") else ""
                next_txt = f"{nxt['label']} ({prefix}{date_fr})"
            else:
                next_txt = "-"
            cumul_txt = f"{p['cumul_dj']}" if p.get("cumul_dj") is not None else "-"
            semis_fr = datetime.strptime(p["date_semis"], "%Y-%m-%d").strftime("%d/%m/%Y")
            pdf.cell(widths2[0], 6, safe(p["nom_parcelle"]), border=1)
            pdf.cell(widths2[1], 6, safe(p["semence"]), border=1)
            pdf.cell(widths2[2], 6, safe(semis_fr), border=1, align="C")
            pdf.cell(widths2[3], 6, safe(cumul_txt), border=1, align="C")
            pdf.cell(widths2[4], 6, safe(stage_label), border=1)
            pdf.cell(widths2[5], 6, safe(next_txt), border=1)
            pdf.ln()

    _add_phenologie_table("Phenologie du mais", pheno_mais or [])
    _add_phenologie_table("Phenologie des cereales a paille", pheno_cereales or [])

    # Anomalies détectées
    if anomalies:
        pdf.ln(4)
        pdf.set_font("Arial", "B", 11)
        pdf.set_fill_color(153, 27, 27); pdf.set_text_color(255, 255, 255)
        pdf.cell(page_w, 8, safe(f"  Anomalies detectees ({len(anomalies)})"), ln=1, fill=True)
        pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 8)
        for a in anomalies[:30]:  # limite raisonnable pour ne pas generer un PDF demesure
            pdf.set_fill_color(254, 242, 242)
            pdf.multi_cell(page_w, 6, safe(f"  [{a.get('type','')}] {a.get('nom_parcelle','')} : {a.get('message','')}"), border=1, fill=True)
        if len(anomalies) > 30:
            pdf.set_font("Arial", "I", 8)
            pdf.cell(page_w, 6, safe(f"  ... et {len(anomalies) - 30} autre(s), voir la page Synthese de campagne."))
            pdf.ln()

    pdf.set_auto_page_break(auto=False)
    pdf.set_y(-15)
    pdf.set_font("Arial", "I", 8)
    pdf.cell(0, 5, safe(f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - Dashboard Agricole v{DASHBOARD_VERSION}"), align="R")

    os.makedirs('exports', exist_ok=True)
    path = os.path.join('exports', f"synthese_campagne_{data['date'].replace('-','')}.pdf")
    pdf.output(path)
    return path


# ================= PHENOLOGIE MAIS (degres-jours base 6-30) =================
# Seuils indicatifs des stades précoces (peu variables entre variétés, contrairement à la
# floraison et la maturité qui dépendent fortement de la précocité de l'hybride -- ceux-ci
# sont donc renseignés par variété dans le catalogue produits, cf. besoin_floraison/maturite).
# Les stades floraison_femelle, grain_laiteux, lentille_vitreuse et recolte sont dérivés par
# décalage/interpolation à partir des besoins variétaux (semis->floraison, semis->maturité) :
# ce sont des ordres de grandeur usuels (Arvalis), pas des valeurs mesurées pour votre variété
# précise -- à affiner avec l'observation terrain, en particulier pour la date de récolte qui
# dépend surtout de la météo de ressuyage/séchage du grain, pas seulement des degrés-jours.
CORN_STAGE_LABELS = [
    ("semis",             "Semis",              "🌰", 0),
    ("emergence",         "Émergence",          "🌱", 80),
    ("3_feuilles",        "3 feuilles",         "🌿", 180),
    ("5_6_feuilles",      "5-6 feuilles",       "🍃", 300),
    ("10_12_feuilles",    "10-12 feuilles",     "🎋", 550),
    ("floraison_male",    "Floraison mâle",     "🌾", None),  # seuil variétal (besoin_floraison)
    ("floraison_femelle", "Floraison femelle",  "🎀", None),  # besoin_floraison + ~75°C.j (soies)
    ("grain_laiteux",     "Grain laiteux",      "🥛", None),  # interpolé entre floraison femelle et maturité
    ("lentille_vitreuse", "Lentille vitreuse",  "💎", None),  # besoin_maturite - ~150°C.j (repère ensilage)
    ("maturite",          "Maturité",           "⚫", None),  # seuil variétal (besoin_maturite, "point noir")
    ("recolte",           "Récolte",            "🌽", None),  # besoin_maturite + ~100°C.j (approximatif, dépend du séchage)
]

def _corn_derived_thresholds(besoin_floraison, besoin_maturite):
    """Calcule les seuils des stades dérivés (non saisis directement) à partir des deux
    besoins variétaux renseignés dans le catalogue (semis->floraison, semis->maturité)."""
    floraison_femelle = (besoin_floraison + 75) if besoin_floraison else None
    grain_laiteux = None
    if floraison_femelle and besoin_maturite and besoin_maturite > floraison_femelle:
        grain_laiteux = round(floraison_femelle + 0.35 * (besoin_maturite - floraison_femelle))
    lentille_vitreuse = round(besoin_maturite - 150) if besoin_maturite else None
    recolte = round(besoin_maturite + 100) if besoin_maturite else None
    return {
        "floraison_male": besoin_floraison,
        "floraison_femelle": floraison_femelle,
        "grain_laiteux": grain_laiteux,
        "lentille_vitreuse": lentille_vitreuse,
        "maturite": besoin_maturite,
        "recolte": recolte,
    }

_dj_cache = {}
_DJ_CACHE_TTL = 3 * 3600  # 3h : la meteo archive ne change pas plus souvent que ca
_last_dj_cache_cleanup = time.time()


def _cleanup_dj_cache(now):
    """
    _dj_cache accumulait indéfiniment une entrée par (parcelle, date de semis, base,
    plafond) jamais retirée une fois expirée -- seule la LECTURE ignorait les entrées trop
    anciennes, sans jamais les supprimer. Sur un serveur qui tourne plusieurs campagnes,
    ça grossit doucement mais sans fin. Balayage global, au plus une fois toutes les 30
    minutes pour rester léger (le TTL lui-même est de 3h, pas la peine de vérifier plus
    souvent).
    """
    global _last_dj_cache_cleanup
    if now - _last_dj_cache_cleanup < 1800:
        return
    _last_dj_cache_cleanup = now
    expired_keys = [k for k, (_, ts) in _dj_cache.items() if now - ts >= _DJ_CACHE_TTL]
    for k in expired_keys:
        _dj_cache.pop(k, None)

def _fetch_historical_daily_temps(lat, lon, start_date, end_date):
    """Retourne une liste de (date, tmax, tmin) via l'API archive (historique) Open-Meteo."""
    try:
        r = requests.get(
            "https://archive-api.open-meteo.com/v1/archive",
            params={
                "latitude": lat, "longitude": lon,
                "start_date": start_date, "end_date": end_date,
                "daily": "temperature_2m_max,temperature_2m_min",
                "timezone": "Europe/Paris",
            },
            timeout=15,
        )
        r.raise_for_status()
        daily = r.json().get("daily", {})
        return list(zip(daily.get("time", []), daily.get("temperature_2m_max", []), daily.get("temperature_2m_min", [])))
    except Exception:
        return []

def _compute_dj_timeline_cached(lat, lon, start_date_str, base_temp=6.0, cap_temp=30.0):
    """
    Série journalière du cumul de degrés-jours (base/plafond paramétrables -- 6-30°C pour le
    maïs, 0°C sans plafond pour les céréales à paille) depuis start_date_str jusqu'à
    aujourd'hui, avec le taux moyen des 14 derniers jours (pour projeter une date estimée
    pour les stades pas encore atteints). Mis en cache (clé arrondie) pour éviter de spammer
    l'API météo à chaque rafraîchissement.
    Retourne {"timeline": [(date, cumul_ce_jour), ...], "cumul_final": float, "taux_moyen_j": float}.
    """
    key = (round(lat, 3), round(lon, 3), start_date_str, base_temp, cap_temp)
    now = time.time()
    _cleanup_dj_cache(now)
    cached = _dj_cache.get(key)
    if cached and (now - cached[1] < _DJ_CACHE_TTL):
        return cached[0]

    end_date_str = datetime.now().strftime("%Y-%m-%d")
    rows = _fetch_historical_daily_temps(lat, lon, start_date_str, end_date_str)
    timeline = []
    cumul = 0.0
    for date_str, tmax, tmin in rows:
        if tmax is None or tmin is None:
            continue
        tmax_ajuste = min(tmax, cap_temp) if cap_temp is not None else tmax
        dj = ((tmax_ajuste + tmin) / 2.0) - base_temp
        if dj > 0:
            cumul += dj
        timeline.append((date_str, round(cumul, 1)))

    # Taux moyen des 14 derniers jours (ou de toute la période si plus courte), pour projeter
    # une date estimée sur les stades pas encore atteints.
    taux_moyen_j = 0.0
    if len(timeline) >= 2:
        fenetre = min(14, len(timeline) - 1)
        cumul_debut = timeline[-1 - fenetre][1]
        cumul_fin = timeline[-1][1]
        taux_moyen_j = round((cumul_fin - cumul_debut) / fenetre, 2) if fenetre > 0 else 0.0

    result = {"timeline": timeline, "cumul_final": timeline[-1][1] if timeline else 0.0, "taux_moyen_j": taux_moyen_j}
    _dj_cache[key] = (result, now)
    return result

def _geofence_centroid(geom):
    if geom["type"] == "circle":
        return geom["center"][0], geom["center"][1]
    coords = geom["coords"]
    return sum(c[0] for c in coords) / len(coords), sum(c[1] for c in coords) / len(coords)

@app.route("/api/phenologie_mais")
@login_or_share_token_required
def api_phenologie_mais():
    return jsonify(_compute_phenologie_mais())


def _compute_phenologie_mais():
    """
    Frise phénologique du maïs, par parcelle (et par sous-parcelle le cas échéant), basée
    sur le cumul de degrés-jours (base 6°C - plafond 30°C, méthode française standard)
    depuis la date de semis. Ne suit que les parcelles semées avec une variété dont les
    besoins semis->floraison et/ou semis->maturité ont été renseignés dans le catalogue
    produits (type "semence") -- fonctionnalité opt-in, sans nouveau champ "culture" sur
    les parcelles.
    """
    DB_PATH = 'database.db'
    _ensure_sous_parcelles_table()
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT name, besoin_floraison, besoin_maturite FROM catalog_products
            WHERE type = 'semence' AND (besoin_floraison IS NOT NULL OR besoin_maturite IS NOT NULL)
        """)
        semences_suivies = {r["name"]: {"besoin_floraison": r["besoin_floraison"], "besoin_maturite": r["besoin_maturite"]} for r in cur.fetchall()}

        if not semences_suivies:
            return {"parcelles": [], "message": "Aucune semence n'a de besoins en degrés-jours renseignés (à saisir dans le catalogue produits, type Semence)."}

        cur.execute("SELECT id, nom, polygon FROM sous_parcelles")
        sous_parcelles_info = {r["id"]: dict(r) for r in cur.fetchall()}

        # Toutes les interventions "Semis" de la campagne en cours utilisant une de ces
        # semences, pour chaque parcelle -- une même parcelle peut porter plusieurs variétés
        # différentes (sous-parcelles distinctes) : chacune est suivie séparément, MÊME si
        # elles utilisent la même semence (le regroupement inclut désormais sous_parcelle_id,
        # pas seulement geofence_id+semence -- sans quoi deux sous-parcelles de la même
        # variété auraient été fusionnées en un seul suivi). En revanche, si la MÊME semence
        # est ressemée deux fois sur la MÊME sous-parcelle (ex: resemis après échec), seule
        # la première occurrence est retenue (comportement inchangé pour ce cas précis).
        cur.execute("SELECT geofence_id, exit_time, products, sous_parcelle_id FROM interventions WHERE intervention_type = 'Semis' ORDER BY exit_time ASC")
        semis_events = []
        seen_pairs = set()
        for row in cur.fetchall():
            try:
                products = json.loads(row["products"] or "[]")
            except Exception:
                products = []
            semence_utilisee = next((p.get("name") for p in products if p.get("type") == "semence" and p.get("name") in semences_suivies), None)
            geo_id_str = str(row["geofence_id"])
            sp_id = row["sous_parcelle_id"]
            if semence_utilisee:
                pair = (geo_id_str, sp_id, semence_utilisee)
                if pair not in seen_pairs:
                    seen_pairs.add(pair)
                    semis_events.append({"geofence_id": geo_id_str, "sous_parcelle_id": sp_id, "date": str(row["exit_time"])[:10], "semence": semence_utilisee})

    if not semis_events:
        return {"parcelles": [], "message": "Aucun semis avec une semence suivie n'a été trouvé pour la campagne en cours."}

    raw = build_data()
    geofences_named = raw.get("geofences", {})

    raw_geofences_list = safe_get(f"{TRACCAR_URL}/geofences")
    parsed_geofences = {}
    if isinstance(raw_geofences_list, list):
        for g in raw_geofences_list:
            geom = _parse_geofence_wkt(g.get("area"))
            if geom:
                parsed_geofences[str(g["id"])] = geom

    result = []
    for info in semis_events:
        geo_id_str = info["geofence_id"]
        sp_id = info.get("sous_parcelle_id")
        sp_info = sous_parcelles_info.get(sp_id) if sp_id else None

        # Géométrie de référence pour le calcul météo : priorité au polygone de la
        # sous-parcelle elle-même (plus précis) quand l'intervention y est rattachée,
        # repli sur la géofence Traccar entière sinon.
        geom = None
        if sp_info and sp_info.get("polygon"):
            try:
                sp_coords = json.loads(sp_info["polygon"])
                if sp_coords and len(sp_coords) >= 3:
                    geom = {"type": "polygon", "coords": sp_coords}
            except Exception:
                geom = None
        if geom is None:
            geom = parsed_geofences.get(geo_id_str)

        semence_info = semences_suivies[info["semence"]]
        besoin_floraison = semence_info.get("besoin_floraison")
        besoin_maturite = semence_info.get("besoin_maturite")
        derives = _corn_derived_thresholds(besoin_floraison, besoin_maturite)

        stages = []
        for key, label, icon, seuil_fixe in CORN_STAGE_LABELS:
            seuil = derives[key] if key in derives else seuil_fixe
            stages.append({"key": key, "label": label, "icon": icon, "seuil": seuil})

        cumul_dj = None
        if geom:
            lat, lon = _geofence_centroid(geom)
            dj_data = _compute_dj_timeline_cached(lat, lon, info["date"])
            timeline = dj_data["timeline"]
            cumul_dj = dj_data["cumul_final"]
            taux_moyen_j = dj_data["taux_moyen_j"]

            # Date de franchissement de chaque seuil : date réelle si déjà atteint (recherche
            # dans l'historique jour par jour), sinon date estimée par extrapolation au taux
            # moyen des 14 derniers jours (approximation, précisée comme telle côté interface).
            for s in stages:
                if s["seuil"] is None:
                    s["date"] = None
                    s["date_estimee"] = False
                    continue
                if s["seuil"] <= 0:
                    s["date"] = info["date"]
                    s["date_estimee"] = False
                    continue
                atteint_le = next((d for d, c in timeline if c >= s["seuil"]), None)
                if atteint_le:
                    s["date"] = atteint_le
                    s["date_estimee"] = False
                elif taux_moyen_j > 0:
                    jours_restants = (s["seuil"] - cumul_dj) / taux_moyen_j
                    s["date"] = (datetime.now() + timedelta(days=jours_restants)).strftime("%Y-%m-%d")
                    s["date_estimee"] = True
                else:
                    s["date"] = None
                    s["date_estimee"] = False
        else:
            for s in stages:
                s["date"] = None
                s["date_estimee"] = False

        stage_atteint = "semis"
        if cumul_dj is not None:
            for s in stages:
                if s["seuil"] is not None and cumul_dj >= s["seuil"]:
                    stage_atteint = s["key"]

        nom_parcelle = resolve_geo_name(geofences_named, geo_id_str)
        if sp_info:
            nom_parcelle = f"{nom_parcelle} — {sp_info['nom']}"

        result.append({
            "geofence_id": geo_id_str,
            "sous_parcelle_id": sp_id,
            "nom_parcelle": nom_parcelle,
            "semence": info["semence"],
            "date_semis": info["date"],
            "cumul_dj": cumul_dj,
            "stage_atteint": stage_atteint,
            "stages": stages,
            "position_connue": geom is not None,
        })

    result.sort(key=lambda p: p["nom_parcelle"].lower())
    return {"parcelles": result}


# ================= PHENOLOGIE CEREALES A PAILLE (degres-jours base 0) =================
# Blé, orge, avoine et triticale partagent la même échelle BBCH officielle (Witzenberger et
# al. 1989) et une température de base identique (0°C, sans plafond documenté à ce jour,
# contrairement au maïs). Modèle générique (pas de saisie par variété comme pour le maïs) :
# seul le stade "épi 1 cm" (~1000°C.j base 0) est solidement sourcé (Arvalis) ; les autres
# seuils sont des ordres de grandeur usuels, à affiner par l'observation terrain au fil des
# campagnes -- comme pour les stades dérivés du maïs.
CEREAL_STAGE_LABELS = [
    ("semis",             "Semis",                  "🌰", 0),
    ("levee",             "Levée",                  "🌱", 150),
    ("tallage",           "Début tallage",          "🌿", 280),
    ("epi_1cm",           "Épi 1 cm",               "📏", 1000),  # seuil sourcé (Arvalis)
    ("derniere_feuille",  "Dernière feuille étalée","🍃", 1450),
    ("epiaison",          "Épiaison",               "🌾", 1650),
    ("floraison",         "Floraison",              "🌸", 1800),
    ("grain_laiteux",     "Grain laiteux",          "🥛", 2050),
    ("maturite",          "Maturité",               "🟤", 2500),
    ("recolte",           "Récolte",                "🚜", 2600),
]

# Mots-clés (insensibles à la casse) permettant de reconnaître une semence de céréale à
# paille via le champ "culture" du catalogue produits, sans dépendre du libellé exact utilisé
# dans votre table de cultures (ex: "Blé tendre d'hiver", "Orge de printemps"...).
STRAW_CEREAL_KEYWORDS = ["blé", "ble", "orge", "avoine", "triticale", "seigle"]

@app.route("/api/phenologie_cereales")
@login_or_share_token_required
def api_phenologie_cereales():
    return jsonify(_compute_phenologie_cereales())


def _compute_phenologie_cereales():
    """
    Frise phénologique des céréales à paille (blé/orge/avoine/triticale), par parcelle,
    basée sur le cumul de degrés-jours (base 0°C, sans plafond) depuis la date de semis.
    Modèle générique commun aux 4 céréales (même échelle BBCH), contrairement au maïs qui
    nécessite une saisie par variété -- ici la détection se fait via le champ "culture" du
    produit semence utilisé au semis.
    """
    DB_PATH = 'database.db'
    _ensure_sous_parcelles_table()
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("SELECT name, culture FROM catalog_products WHERE type = 'semence'")
        semences_cereales = set()
        for r in cur.fetchall():
            culture_lower = (r["culture"] or "").lower()
            if any(kw in culture_lower for kw in STRAW_CEREAL_KEYWORDS):
                semences_cereales.add(r["name"])

        if not semences_cereales:
            return {"parcelles": [], "message": "Aucune semence de céréale à paille (blé/orge/avoine/triticale) trouvée -- renseignez le champ Culture sur la fiche semence dans le catalogue produits."}

        cur.execute("SELECT id, nom, polygon FROM sous_parcelles")
        sous_parcelles_info = {r["id"]: dict(r) for r in cur.fetchall()}

        # Regroupement par (géofence, sous-parcelle, semence) : deux sous-parcelles de la
        # même géofence portant la même semence sont désormais suivies séparément (avant,
        # elles auraient été fusionnées en un seul suivi, faussant le stade affiché pour
        # l'une des deux -- voir la même correction sur le maïs pour le détail).
        cur.execute("SELECT geofence_id, exit_time, products, sous_parcelle_id FROM interventions WHERE intervention_type = 'Semis' ORDER BY exit_time ASC")
        semis_events = []
        seen_pairs = set()
        for row in cur.fetchall():
            try:
                products = json.loads(row["products"] or "[]")
            except Exception:
                products = []
            semence_utilisee = next((p.get("name") for p in products if p.get("type") == "semence" and p.get("name") in semences_cereales), None)
            geo_id_str = str(row["geofence_id"])
            sp_id = row["sous_parcelle_id"]
            if semence_utilisee:
                pair = (geo_id_str, sp_id, semence_utilisee)
                if pair not in seen_pairs:
                    seen_pairs.add(pair)
                    semis_events.append({"geofence_id": geo_id_str, "sous_parcelle_id": sp_id, "date": str(row["exit_time"])[:10], "semence": semence_utilisee})

    if not semis_events:
        return {"parcelles": [], "message": "Aucun semis de céréale à paille n'a été trouvé pour la campagne en cours."}

    raw = build_data()
    geofences_named = raw.get("geofences", {})

    raw_geofences_list = safe_get(f"{TRACCAR_URL}/geofences")
    parsed_geofences = {}
    if isinstance(raw_geofences_list, list):
        for g in raw_geofences_list:
            geom = _parse_geofence_wkt(g.get("area"))
            if geom:
                parsed_geofences[str(g["id"])] = geom

    result = []
    for info in semis_events:
        geo_id_str = info["geofence_id"]
        sp_id = info.get("sous_parcelle_id")
        sp_info = sous_parcelles_info.get(sp_id) if sp_id else None

        # Géométrie de référence : priorité au polygone de la sous-parcelle (plus précis),
        # repli sur la géofence Traccar entière sinon.
        geom = None
        if sp_info and sp_info.get("polygon"):
            try:
                sp_coords = json.loads(sp_info["polygon"])
                if sp_coords and len(sp_coords) >= 3:
                    geom = {"type": "polygon", "coords": sp_coords}
            except Exception:
                geom = None
        if geom is None:
            geom = parsed_geofences.get(geo_id_str)

        stages = [{"key": key, "label": label, "icon": icon, "seuil": seuil} for key, label, icon, seuil in CEREAL_STAGE_LABELS]

        cumul_dj = None
        if geom:
            lat, lon = _geofence_centroid(geom)
            dj_data = _compute_dj_timeline_cached(lat, lon, info["date"], base_temp=0.0, cap_temp=None)
            timeline = dj_data["timeline"]
            cumul_dj = dj_data["cumul_final"]
            taux_moyen_j = dj_data["taux_moyen_j"]

            for s in stages:
                if s["seuil"] <= 0:
                    s["date"] = info["date"]
                    s["date_estimee"] = False
                    continue
                atteint_le = next((d for d, c in timeline if c >= s["seuil"]), None)
                if atteint_le:
                    s["date"] = atteint_le
                    s["date_estimee"] = False
                elif taux_moyen_j > 0:
                    jours_restants = (s["seuil"] - cumul_dj) / taux_moyen_j
                    s["date"] = (datetime.now() + timedelta(days=jours_restants)).strftime("%Y-%m-%d")
                    s["date_estimee"] = True
                else:
                    s["date"] = None
                    s["date_estimee"] = False
        else:
            for s in stages:
                s["date"] = None
                s["date_estimee"] = False

        stage_atteint = "semis"
        if cumul_dj is not None:
            for s in stages:
                if cumul_dj >= s["seuil"]:
                    stage_atteint = s["key"]

        nom_parcelle = resolve_geo_name(geofences_named, geo_id_str)
        if sp_info:
            nom_parcelle = f"{nom_parcelle} — {sp_info['nom']}"

        result.append({
            "geofence_id": geo_id_str,
            "sous_parcelle_id": sp_id,
            "nom_parcelle": nom_parcelle,
            "semence": info["semence"],
            "date_semis": info["date"],
            "cumul_dj": cumul_dj,
            "stage_atteint": stage_atteint,
            "stages": stages,
            "position_connue": geom is not None,
        })

    result.sort(key=lambda p: p["nom_parcelle"].lower())
    return {"parcelles": result}



@app.route("/api/incoherences")
@login_or_share_token_required
def api_incoherences():
    """
    Détecte des anomalies de données (chronologies impossibles, surfaces incohérentes,
    doublons de saisie suspects) sur l'ensemble du carnet -- pour repérer des erreurs de
    saisie plutôt que de les découvrir a posteriori. Ce sont des signaux "à vérifier",
    pas des erreurs bloquantes : certains cas peuvent être légitimes.
    """
    DB_PATH = 'database.db'
    _ensure_sous_parcelles_table()
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT device_id, geofence_id, intervention_type, exit_time, applied_area, sous_parcelle_id
            FROM interventions ORDER BY geofence_id, exit_time ASC
        """)
        interventions = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT geofence_id, surface_ha FROM parcelles WHERE surface_ha IS NOT NULL AND surface_ha > 0")
        surfaces_cadastrales = {r["geofence_id"]: r["surface_ha"] for r in cur.fetchall()}
        cur.execute("SELECT id, nom, surface_ha FROM sous_parcelles")
        sous_parcelles_info = {r["id"]: dict(r) for r in cur.fetchall()}

    raw = build_data()
    geofences = raw.get("geofences", {})

    def _parse_dt(s):
        try:
            return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d")
            except Exception:
                return None

    anomalies = []

    # Regroupe par parcelle ET sous-parcelle pour les vérifications de chronologie et de
    # doublons : sans ça, un semis sur une sous-parcelle suivi d'une récolte sur une AUTRE
    # sous-parcelle de la même géofence pourrait déclencher une fausse alerte de chronologie,
    # et deux interventions de sous-parcelles différentes proches dans le temps pourraient se
    # faire passer à tort pour un doublon.
    par_groupe = {}
    for iv in interventions:
        key = (iv["geofence_id"], iv.get("sous_parcelle_id"))
        par_groupe.setdefault(key, []).append(iv)

    for (geo_id, sp_id), ivs in par_groupe.items():
        nom = resolve_geo_name(geofences, geo_id)
        sp_info = sous_parcelles_info.get(sp_id) if sp_id else None
        if sp_info:
            nom = f"{nom} — {sp_info['nom']}"
        ivs_dt = [(iv, _parse_dt(iv["exit_time"])) for iv in ivs]
        ivs_dt = [(iv, dt) for iv, dt in ivs_dt if dt is not None]

        # 1) Chronologie suspecte : un Semis moins de 60 jours après une Récolte sur la
        #    même parcelle (un nouveau cycle démarre rarement aussi vite après la récolte
        #    précédente -- signal à vérifier, pas une impossibilité absolue).
        recoltes = [(iv, dt) for iv, dt in ivs_dt if iv["intervention_type"] == "Récolte"]
        semis = [(iv, dt) for iv, dt in ivs_dt if iv["intervention_type"] == "Semis"]
        for r_iv, r_dt in recoltes:
            for s_iv, s_dt in semis:
                delta = (s_dt - r_dt).days
                if 0 <= delta < 60:
                    anomalies.append({
                        "type": "chronologie",
                        "geofence_id": geo_id, "nom_parcelle": nom,
                        "message": f"Semis du {s_dt.strftime('%d/%m/%Y')} seulement {delta} j après la récolte du {r_dt.strftime('%d/%m/%Y')} -- à vérifier.",
                    })

        # 2) Doublon suspect : deux interventions du même type, même véhicule, à moins de
        #    2h d'écart sur la même parcelle -- signe possible d'une double saisie.
        ivs_sorted = sorted(ivs_dt, key=lambda x: x[1])
        for i in range(len(ivs_sorted) - 1):
            iv1, dt1 = ivs_sorted[i]
            iv2, dt2 = ivs_sorted[i + 1]
            if (iv1["intervention_type"] == iv2["intervention_type"]
                    and iv1["device_id"] == iv2["device_id"]
                    and 0 <= (dt2 - dt1).total_seconds() <= 7200):
                anomalies.append({
                    "type": "doublon",
                    "geofence_id": geo_id, "nom_parcelle": nom,
                    "message": f"Deux interventions \"{iv1['intervention_type']}\" à {dt1.strftime('%d/%m %H:%M')} et {dt2.strftime('%d/%m %H:%M')} (moins de 2h d'écart, même véhicule) -- doublon possible.",
                })

        # 3) Surface incohérente : surface saisie très différente de la surface de
        #    référence -- celle de la SOUS-PARCELLE si l'intervention y est rattachée
        #    (forcément plus petite que la parcelle entière), sinon la surface cadastrale
        #    de la parcelle entière. Comparer systématiquement à la surface cadastrale
        #    complète, même pour une intervention rattachée à une sous-parcelle plus
        #    petite, déclenchait une fausse alerte "surface trop petite" à chaque saisie.
        surf_ref = (sp_info.get("surface_ha") if sp_info else None) or surfaces_cadastrales.get(geo_id)
        if surf_ref:
            for iv, dt in ivs_dt:
                if iv["applied_area"] and iv["applied_area"] > 0:
                    ratio = iv["applied_area"] / surf_ref
                    if ratio > 1.5 or ratio < 0.15:
                        anomalies.append({
                            "type": "surface",
                            "geofence_id": geo_id, "nom_parcelle": nom,
                            "message": f"Surface saisie {iv['applied_area']} ha le {dt.strftime('%d/%m/%Y')} très différente de la surface de référence ({surf_ref} ha) -- à vérifier.",
                        })

    return jsonify({"anomalies": anomalies, "nb_anomalies": len(anomalies)})


@app.route("/api/today_summary")
@login_required
def today_summary():
    """
    Résumé condensé pour la page d'accueil : interventions du jour, nombre d'alertes
    DAR actives, parcelles nécessitant une attention (statut 'attente' ou 'préparé'
    depuis longtemps), et position GPS du premier véhicule pour la météo locale.
    """
    DB_PATH = 'database.db'
    today_str = datetime.now().strftime("%Y-%m-%d")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        # Interventions du jour (saisies dans le carnet)
        cur.execute("""
            SELECT device_id, geofence_id, intervention_type, exit_time, products, applied_area
            FROM interventions
            WHERE exit_time LIKE ?
            ORDER BY exit_time DESC
        """, (f"{today_str}%",))
        interventions_today = [dict(r) for r in cur.fetchall()]

        # Parcelles en attente ou préparées, qui mériteraient une action
        cur.execute("""
            SELECT geofence_id, identifiant, nom_parcelle, statut
            FROM parcelles
            WHERE statut IN ('attente', 'prepare')
        """)
        parcelles_attention = [dict(r) for r in cur.fetchall()]

        # Parcelles "actives" (hors récolte) n'ayant reçu aucune intervention depuis
        # longtemps -- signe possible d'un oubli ou d'un besoin de surveillance/traitement.
        cur.execute("""
            SELECT geofence_id, identifiant, nom_parcelle, statut
            FROM parcelles
            WHERE statut IS NULL OR statut NOT IN ('recolte')
        """)
        parcelles_suivies = [dict(r) for r in cur.fetchall()]

        # Produits dont le stock restant est descendu au niveau (ou en dessous) du seuil
        # d'alerte configuré sur la fiche produit -- évite de découvrir la rupture au
        # moment de vouloir traiter.
        produits_stock_faible = []
        try:
            cur.execute("""
                SELECT name, type, unit, stock, seuil_alerte_stock
                FROM catalog_products
                WHERE stock IS NOT NULL AND seuil_alerte_stock IS NOT NULL
                  AND stock <= seuil_alerte_stock AND (actif IS NULL OR actif = 1)
            """)
            produits_stock_faible = [dict(r) for r in cur.fetchall()]
        except Exception:
            produits_stock_faible = []  # colonnes stock pas encore migrées : ignore silencieusement

        cur.execute("SELECT geofence_id, MAX(exit_time) AS derniere FROM interventions GROUP BY geofence_id")
        dernieres_interventions = {row["geofence_id"]: row["derniere"] for row in cur.fetchall()}

    parcelles_inactives = []
    now = datetime.now()
    for p in parcelles_suivies:
        derniere = dernieres_interventions.get(p["geofence_id"])
        jours = None
        if derniere:
            try:
                jours = (now - datetime.strptime(derniere[:19], "%Y-%m-%dT%H:%M:%S")).days
            except Exception:
                jours = None
        if jours is None or jours >= ALERT_JOURS_SANS_INTERVENTION:
            p2 = dict(p)
            p2["jours_sans_intervention"] = jours  # None = jamais travaillée
            parcelles_inactives.append(p2)

    # Réutiliser la logique d'alertes DAR déjà en place
    alertes_response = alertes_dar()
    try:
        alertes_data = alertes_response.get_json()
    except Exception:
        alertes_data = {"alertes": [], "nb_alertes": 0}

    # Position GPS du premier véhicule actif, pour que le front affiche la météo locale
    raw = build_data()
    first_position = None
    for p in raw.get("positions", []):
        if p.get("lat") and p.get("lon"):
            first_position = {"lat": p["lat"], "lon": p["lon"]}
            break

    # Repli : interroge l'endpoint Traccar des positions "live" si build_data() (historique
    # des 30 derniers jours) n'a rien trouvé -- voir /api/synthese_campagne pour le détail.
    if not first_position:
        try:
            live_positions = safe_get(f"{TRACCAR_URL}/positions")
            if isinstance(live_positions, list):
                for p in live_positions:
                    if p.get("latitude") and p.get("longitude"):
                        first_position = {"lat": p["latitude"], "lon": p["longitude"]}
                        break
        except Exception:
            pass

    geofences = raw.get("geofences", {})
    for p in parcelles_attention:
        p['nom_traccar'] = resolve_geo_name(geofences, p['geofence_id'])
    for p in parcelles_inactives:
        p['nom_traccar'] = resolve_geo_name(geofences, p['geofence_id'])

    for interv in interventions_today:
        interv['nom_parcelle'] = resolve_geo_name(geofences, interv['geofence_id'])

    return jsonify({
        "date": today_str,
        "interventions_today": interventions_today,
        "nb_interventions_today": len(interventions_today),
        "parcelles_attention": parcelles_attention,
        "nb_parcelles_attention": len(parcelles_attention),
        "parcelles_inactives": parcelles_inactives,
        "nb_parcelles_inactives": len(parcelles_inactives),
        "seuil_jours_sans_intervention": ALERT_JOURS_SANS_INTERVENTION,
        "alertes_dar": alertes_data.get("alertes", []),
        "nb_alertes_dar": alertes_data.get("nb_alertes", 0),
        "produits_stock_faible": produits_stock_faible,
        "nb_produits_stock_faible": len(produits_stock_faible),
        "password_is_default": (LOGIN_PASSWORD == _DEFAULT_LOGIN_PASSWORD),
        "first_position": first_position,
    })


@app.route("/api/alertes_dar")
@login_required
def alertes_dar():
    """
    Vérifie la cohérence DAR/récolte : pour chaque récolte enregistrée, vérifie si un
    produit phyto utilisé avant cette récolte sur la même parcelle avait encore un
    DAR (Délai Avant Récolte) actif au moment de la récolte. Remonte les anomalies.
    """
    DB_PATH = 'database.db'
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT geofence_id, intervention_type, exit_time, products
            FROM interventions
            ORDER BY exit_time ASC
        """)
        all_interventions = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT name, type, dar FROM catalog_products WHERE type = 'phyto' AND dar > 0")
        catalog_dar = {r['name'].strip(): r['dar'] for r in cur.fetchall()}

    raw = build_data()
    geofences = raw.get("geofences", {})

    alertes = []
    recoltes = [i for i in all_interventions if i['intervention_type'] == 'Récolte']

    for recolte in recoltes:
        try:
            dt_recolte = datetime.strptime(recolte['exit_time'][:19], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            try:
                dt_recolte = datetime.strptime(recolte['exit_time'][:10], "%Y-%m-%d")
            except Exception:
                continue

        # Chercher les interventions de pulvérisation AVANT cette récolte, sur la même parcelle
        for interv in all_interventions:
            if interv['geofence_id'] != recolte['geofence_id']:
                continue
            if interv['intervention_type'] not in ('Pulvérisation', 'Épandage'):
                continue
            try:
                dt_traitement = datetime.strptime(interv['exit_time'][:19], "%Y-%m-%dT%H:%M:%S")
            except Exception:
                try:
                    dt_traitement = datetime.strptime(interv['exit_time'][:10], "%Y-%m-%d")
                except Exception:
                    continue

            if dt_traitement >= dt_recolte:
                continue  # traitement après la récolte, pas pertinent ici

            try:
                products = json.loads(interv.get('products') or '[]')
            except Exception:
                products = []

            for prod in products:
                prod_name = (prod.get('name') or '').strip()
                dar_jours = catalog_dar.get(prod_name)
                if not dar_jours:
                    continue
                jours_ecoules = (dt_recolte - dt_traitement).days
                if jours_ecoules < dar_jours:
                    alertes.append({
                        'geofence_id': recolte['geofence_id'],
                        'nom_parcelle': resolve_geo_name(geofences, recolte['geofence_id']),
                        'produit': prod_name,
                        'date_traitement': dt_traitement.strftime("%d/%m/%Y"),
                        'date_recolte': dt_recolte.strftime("%d/%m/%Y"),
                        'dar_requis_jours': dar_jours,
                        'jours_ecoules': jours_ecoules,
                        'jours_manquants': dar_jours - jours_ecoules,
                    })

    return jsonify({"alertes": alertes, "nb_alertes": len(alertes)})


@app.route("/api/meteo_intervention")
@login_required
def api_meteo_intervention():
    """
    Météo pour une intervention donnée, calculée aux coordonnées RÉELLES de la parcelle (ou
    sous-parcelle) concernée -- pas à la position de la personne qui saisit les données, ni à
    celle d'un véhicule Traccar quelconque. Bien plus fiable que de compter sur la
    géolocalisation du navigateur (souvent refusée, ou simplement dénuée de sens si la saisie
    se fait depuis le bureau plutôt que depuis le champ).
    Paramètres : geofence_id (obligatoire), sous_parcelle_id (optionnel), datetime (optionnel,
    ISO ; si absent ou récent, utilise les prévisions/conditions actuelles).
    """
    geofence_id = request.args.get("geofence_id")
    sous_parcelle_id = request.args.get("sous_parcelle_id")
    dt_str = request.args.get("datetime")

    if not geofence_id:
        return jsonify({"error": "geofence_id requis"}), 400

    lat = lon = None

    # Priorité à la sous-parcelle si elle est précisée (contour plus précis que la parcelle
    # entière, notamment utile quand la parcelle est scindée en plusieurs cultures).
    if sous_parcelle_id:
        try:
            _ensure_sous_parcelles_table()
            with sqlite3.connect('database.db') as conn:
                conn.row_factory = sqlite3.Row
                row = conn.execute("SELECT polygon FROM sous_parcelles WHERE id = ?", (sous_parcelle_id,)).fetchone()
                if row:
                    coords = json.loads(row["polygon"])
                    if coords:
                        lat = sum(c[0] for c in coords) / len(coords)
                        lon = sum(c[1] for c in coords) / len(coords)
        except Exception:
            pass

    # À défaut, la géométrie de la géofence Traccar elle-même.
    if lat is None:
        try:
            raw_geofences_list = safe_get(f"{TRACCAR_URL}/geofences")
            if isinstance(raw_geofences_list, list):
                for g in raw_geofences_list:
                    if str(g.get("id")) == str(geofence_id):
                        geom = _parse_geofence_wkt(g.get("area"))
                        if geom:
                            lat, lon = _geofence_centroid(geom)
                        break
        except Exception:
            pass

    if lat is None:
        return jsonify({"error": "Impossible de localiser cette parcelle (géométrie introuvable)"}), 404

    try:
        now = datetime.now()
        target_dt = datetime.strptime(dt_str[:19], "%Y-%m-%dT%H:%M:%S") if dt_str else now
    except Exception:
        target_dt = now

    is_historical = target_dt < (now - timedelta(hours=24))

    try:
        if is_historical:
            date_str = target_dt.strftime("%Y-%m-%d")
            r = requests.get(
                "https://archive-api.open-meteo.com/v1/archive",
                params={"latitude": lat, "longitude": lon, "start_date": date_str, "end_date": date_str,
                        "hourly": "temperature_2m,windspeed_10m,precipitation,weathercode", "timezone": "Europe/Paris"},
                timeout=15,
            )
            r.raise_for_status()
            data = r.json()
            hourly = data.get("hourly", {})
            times = hourly.get("time", [])
            if not times:
                return jsonify({"error": "Aucune donnée météo disponible pour cette date (archive Open-Meteo : quelques jours de délai pour les dates très récentes)"}), 404
            h = min(target_dt.hour, len(times) - 1)
            result = {
                "temperature_2m": hourly.get("temperature_2m", [None])[h] if h < len(hourly.get("temperature_2m", [])) else None,
                "windspeed_10m": hourly.get("windspeed_10m", [None])[h] if h < len(hourly.get("windspeed_10m", [])) else None,
                "precipitation": hourly.get("precipitation", [None])[h] if h < len(hourly.get("precipitation", [])) else None,
                "weathercode": hourly.get("weathercode", [None])[h] if h < len(hourly.get("weathercode", [])) else None,
            }
        else:
            r = requests.get(
                "https://api.open-meteo.com/v1/forecast",
                params={"latitude": lat, "longitude": lon,
                        "current": "temperature_2m,windspeed_10m,precipitation,weathercode", "timezone": "Europe/Paris"},
                timeout=15,
            )
            r.raise_for_status()
            result = r.json().get("current", {})
    except Exception as e:
        return jsonify({"error": f"Erreur Open-Meteo : {e}"}), 502

    return jsonify(result)


@app.route("/api/config_meteo_ref", methods=["GET", "POST"])
@login_required
def api_config_meteo_ref():
    """
    Position de reference de l'exploitation (lat/lon), utilisee pour les previsions meteo a
    7 jours affichees sur le tableau de bord. Meme pattern de stockage que /api/config_partage
    (config.json), par coherence avec le reste des reglages.
    """
    import json as _json
    config_path = "config.json"
    if request.method == "GET":
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        mcfg = cfg.get("meteo_ref", {})
        return jsonify({"lat": mcfg.get("lat"), "lon": mcfg.get("lon")})
    else:
        data = request.get_json(silent=True) or {}
        try:
            lat = float(data.get("lat"))
            lon = float(data.get("lon"))
        except (TypeError, ValueError):
            return jsonify({"error": "lat/lon invalides"}), 400
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        cfg["meteo_ref"] = {"lat": lat, "lon": lon}
        with open(config_path, "w") as f: _json.dump(cfg, f, indent=2)
        return jsonify({"status": "ok"})


@app.route("/api/meteo_previsions")
@login_required
def api_meteo_previsions():
    """
    Previsions meteo a 4 jours pour le tableau de bord (modele officiel Meteo-France ARPEGE
    Europe, via Open-Meteo) -- utile pour anticiper une fenetre de traitement (vent, pluie a
    venir), contrairement a la meteo historique deja utilisee a la date d'une intervention
    passee (api_meteo_intervention). Horizon limite a 4 jours : c'est la portee reelle du
    modele ARPEGE Europe, pas une limitation arbitraire de cette route.

    Utilise la position de reference de l'exploitation (config.json, section "meteo_ref"),
    ou lat/lon en parametres de requete pour la surcharger ponctuellement.
    """
    import json as _json
    lat = request.args.get("lat", type=float)
    lon = request.args.get("lon", type=float)
    if lat is None or lon is None:
        cfg = {}
        try:
            if os.path.exists("config.json"):
                with open("config.json", "r") as f: cfg = _json.load(f)
        except Exception: pass
        mcfg = cfg.get("meteo_ref", {})
        lat = mcfg.get("lat")
        lon = mcfg.get("lon")
    if lat is None or lon is None:
        return jsonify({"error": "Position de reference non configuree (voir ⚙️ État système)"}), 400

    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={"latitude": lat, "longitude": lon,
                    "daily": "weathercode,temperature_2m_max,temperature_2m_min,precipitation_sum,windspeed_10m_max",
                    "timezone": "Europe/Paris", "forecast_days": 4,
                    # Modele officiel Meteo-France ARPEGE Europe (resolution ~11km), plutot que
                    # le modele "best match" generique -- prevision plus fiable localement,
                    # mais horizon reel limite a 4 jours (contre 7-16 en mode generique).
                    "models": "meteofrance_arpege_europe"},
            timeout=15,
        )
        r.raise_for_status()
        daily = r.json().get("daily", {})
    except Exception as e:
        return jsonify({"error": f"Erreur Open-Meteo : {e}"}), 502

    jours = []
    dates = daily.get("time", [])
    for i, date_str in enumerate(dates):
        jours.append({
            "date": date_str,
            "temp_min": daily.get("temperature_2m_min", [None]*len(dates))[i],
            "temp_max": daily.get("temperature_2m_max", [None]*len(dates))[i],
            "precipitation": daily.get("precipitation_sum", [None]*len(dates))[i],
            "vent_max": daily.get("windspeed_10m_max", [None]*len(dates))[i],
            "weathercode": daily.get("weathercode", [None]*len(dates))[i],
        })
    return jsonify({"jours": jours})



@app.route("/data")
@login_required
def data():
    vehicle = request.args.get("vehicle","")
    geofence = request.args.get("geofence","")
    start = request.args.get("start")
    end = request.args.get("end")

    # Calculer dynamiquement DAYS_BACK selon la période demandée
    global DAYS_BACK, _cached_data, _last_cache_time
    if start:
        try:
            dt_start = datetime.strptime(start[:16], "%Y-%m-%dT%H:%M")
            days_needed = (datetime.utcnow() - dt_start).days + 1
            if days_needed != DAYS_BACK:
                DAYS_BACK = max(1, days_needed)
                _cached_data = None  # Invalider le cache
                _last_cache_time = 0
        except Exception:
            pass

    raw = build_data()
    events = apply_filters(raw["events"], vehicle, geofence, start, end)
    return jsonify({"events": events, "positions": raw["positions"], "geofences": raw["geofences"]})

# Exports du rapport de chantier (Excel/PDF) : extraits dans chantier_export_bp.py,
# voir l'enregistrement du blueprint en tete de fichier.


# Cahier de tracabilite PDF : extrait dans cahier_bp.py, voir l'enregistrement
# du blueprint en tete de fichier.



@app.route("/export_phyto")
@login_required
def export_phyto():
    """Export XML registre phyto/semis."""
    DB_PATH = 'database.db'

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT device_id, geofence_id, exit_time,
                   intervention_type, products, applied_area, meteo
            FROM interventions
            WHERE intervention_type IN ('Pulvérisation', 'Semis')
            ORDER BY exit_time DESC
        """)
        interventions = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT name, type, amm, unit, culture, bbch, target, bio FROM catalog_products")
        catalog = {r['name']: dict(r) for r in cur.fetchall()}

        cur.execute("SELECT nom, code_oepp FROM cultures")
        cultures_oepp = {r['nom'].strip().lower(): r['code_oepp'] for r in cur.fetchall()}

        cur.execute("SELECT geofence_id, identifiant, surface_ha FROM parcelles")
        parcelles_rows = cur.fetchall()
        parcelles = {str(r['geofence_id']): r['identifiant'] for r in parcelles_rows}
        parcelles_surface = {int(r['geofence_id']): r['surface_ha'] for r in parcelles_rows if r['surface_ha']}

        cur.execute("SELECT siret, raison_sociale, applicateur, certiphyto, materiel, num_controle, date_controle FROM exploitation WHERE id = 1")
        row = cur.fetchone()
        siret = row['siret'] if row else ''
        raison_sociale = row['raison_sociale'] if row else ''
        applicateur = row['applicateur'] if row else ''
        certiphyto = row['certiphyto'] if row else ''
        materiel = row['materiel'] if row else ''
        num_controle = row['num_controle'] if row else ''
        date_controle = row['date_controle'] if row else ''

    raw = build_data()
    gps_index = {}
    for e in raw["events"]:
        if e["type"] == "Sortie":
            key = (str(e["deviceId"]), str(e["geofenceId"]), e["date"][:19])
            gps_index[key] = (e.get("lat", ""), e.get("lon", ""))
    geofences = raw.get("geofences", {})

    def esc(s):
        return str(s or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;").replace("'","&apos;")

    # Le registre phyto ne doit contenir QUE les produits phytosanitaires et semences,
    # jamais les engrais (même si saisis dans une intervention de type Pulvérisation/Épandage)
    TYPES_AUTORISES_REGISTRE = {'phyto', 'semence'}

    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<registre_phytosanitaire>',
             '  <exploitation>',
             f'    <siret>{esc(siret)}</siret>',
             f'    <raison_sociale>{esc(raison_sociale)}</raison_sociale>',
             f'    <applicateur>{esc(applicateur)}</applicateur>',
             f'    <numero_certiphyto>{esc(certiphyto)}</numero_certiphyto>',
             '  </exploitation>',
             '  <materiel_pulverisation>',
             f'    <nom>{esc(materiel)}</nom>',
             f'    <numero_controle>{esc(num_controle)}</numero_controle>',
             f'    <date_dernier_controle>{esc(date_controle)}</date_dernier_controle>',
             '  </materiel_pulverisation>',
             '  <interventions>']

    for interv in interventions:
        geo_id_str = str(interv['geofence_id'])
        geo_name = geofences.get(geo_id_str, {}).get('name', f"Parcelle {interv['geofence_id']}")
        id_parcelle = parcelles.get(geo_id_str, '')
        gps_key = (str(interv['device_id']), geo_id_str, interv['exit_time'][:19])
        lat, lon = gps_index.get(gps_key, ("", ""))
        try:
            date_only = datetime.strptime(interv['exit_time'][:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            date_only = interv['exit_time'][:10]
        geo_id_int = int(interv.get('geofence_id', 0) or 0)
        surf_cadastrale_exp = parcelles_surface.get(geo_id_int)
        if surf_cadastrale_exp:
            surf = f"{surf_cadastrale_exp} ha"
        else:
            surf = (str(interv['applied_area']) + ' ha') if interv.get('applied_area') is not None else ''
        try:
            products = json.loads(interv['products']) if interv['products'] else []
        except Exception:
            products = []

        # Parser météo
        meteo = {}
        try:
            meteo = json.loads(interv.get('meteo') or '{}') or {}
        except Exception:
            meteo = {}

        # Construire la liste des produits éligibles AVANT toute écriture XML.
        # Le registre phyto exclut les engrais : seuls phyto et semences y figurent.
        produits_xml = []
        for prod in products:
            prod_name = prod.get('name', '')
            cat = catalog.get(prod_name, {})
            if cat.get('type') not in TYPES_AUTORISES_REGISTRE:
                continue
            prod_dose = prod.get('dosage', '')
            unit = cat.get('unit', '')
            dose_display = f"{prod_dose} {unit}/ha".strip() if prod_dose else ''
            produits_xml += [
                '        <produit>',
                f'          <nom>{esc(prod_name)}</nom>',
                f'          <numero_amm>{esc(cat.get("amm",""))}</numero_amm>',
                f'          <dose>{esc(dose_display)}</dose>',
                f'          <cible>{esc(cat.get("target",""))}</cible>',
                f'          <code_culture>{esc(cultures_oepp.get((cat.get("culture") or "").strip().lower(), "") or cat.get("culture",""))}</code_culture>',
                f'          <stade_bbch>{esc(cat.get("bbch",""))}</stade_bbch>',
                f'          <bio>{"Oui" if cat.get("bio") else "Non"}</bio>',
                '        </produit>',
            ]

        # Si aucun produit éligible (intervention 100% engrais), on saute entièrement
        # cette intervention : rien n'est ajouté à lines.
        if not produits_xml:
            continue

        lines += [
            '    <intervention>',
            f'      <date>{esc(date_only)}</date>',
            f'      <type>{esc(interv["intervention_type"])}</type>',
            f'      <applicateur>{esc(applicateur)}</applicateur>',
            f'      <numero_certiphyto>{esc(certiphyto)}</numero_certiphyto>',
            '      <meteo>',
            f'        <conditions>{esc(meteo.get("conditions",""))}</conditions>',
            f'        <temperature unite="°C">{esc(meteo.get("temperature",""))}</temperature>',
            f'        <vent unite="km/h">{esc(meteo.get("vent",""))}</vent>',
            f'        <pluie unite="mm">{esc(meteo.get("pluie",""))}</pluie>',
            '      </meteo>',
            '      <parcelle>',
            f'        <identifiant>{esc(id_parcelle)}</identifiant>',
            f'        <nom>{esc(id_parcelle if id_parcelle else geo_name)}</nom>',
            f'        <surface_travaillee unite="ha">{esc(surf)}</surface_travaillee>',
            f'        <gps_lat>{esc(lat)}</gps_lat>',
            f'        <gps_lon>{esc(lon)}</gps_lon>',
            '      </parcelle>',
            '      <produits>',
        ]
        lines += produits_xml
        lines += ['      </produits>', '    </intervention>']

    lines += ['  </interventions>', '</registre_phytosanitaire>']

    xml_content = "\n".join(lines)
    return send_file(
        io.BytesIO(xml_content.encode('utf-8')),
        as_attachment=True,
        download_name="registre_phyto.xml",
        mimetype='application/xml'
    )


# Sauvegarde/restauration de la configuration applicative : extrait dans
# config_backup_bp.py, voir l'enregistrement du blueprint en tete de fichier.



@app.route("/api/system_status")
@login_required
def system_status():
    """État du système : BDD, sauvegardes, dernière synchro Traccar."""
    DB_PATH = 'database.db'
    status = {'dashboard_version': DASHBOARD_VERSION}

    # Taille et dernière modif de la BDD
    if os.path.exists(DB_PATH):
        status['db_size_kb'] = round(os.path.getsize(DB_PATH) / 1024, 1)
        status['db_last_modified'] = datetime.fromtimestamp(os.path.getmtime(DB_PATH)).strftime("%d/%m/%Y %Hh%M")
    else:
        status['db_size_kb'] = 0
        status['db_last_modified'] = 'N/A'

    # Nombre d'interventions et de produits en BDD
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM interventions")
            status['nb_interventions'] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM catalog_products")
            status['nb_produits'] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM parcelles WHERE identifiant != ''")
            status['nb_parcelles_identifiees'] = cur.fetchone()[0]
    except Exception as e:
        status['db_error'] = str(e)

    # Liste des sauvegardes disponibles
    backup_dir = "backups"
    backups = []
    if os.path.exists(backup_dir):
        for fname in sorted(os.listdir(backup_dir), reverse=True):
            fpath = os.path.join(backup_dir, fname)
            if os.path.isfile(fpath):
                backups.append({
                    'name': fname,
                    'size_kb': round(os.path.getsize(fpath) / 1024, 1),
                    'date': datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%d/%m/%Y %Hh%M")
                })
    status['backups'] = backups
    status['backups_total_kb'] = round(sum(b['size_kb'] for b in backups), 1)

    # Dernière synchro Traccar (basée sur le cache)
    global _last_cache_time
    if _last_cache_time:
        status['last_traccar_sync'] = datetime.fromtimestamp(_last_cache_time).strftime("%d/%m/%Y %Hh%M")
        status['cache_age_seconds'] = round(time.time() - _last_cache_time)
    else:
        status['last_traccar_sync'] = 'Jamais'
        status['cache_age_seconds'] = None

    # Test de connexion Traccar en direct
    try:
        test = safe_get(f"{TRACCAR_URL}/devices")
        status['traccar_reachable'] = bool(test)
    except Exception:
        status['traccar_reachable'] = False

    status['traccar_url']       = TRACCAR_URL
    status['traccar_user']      = TRACCAR_USER
    status['traccar_days_back'] = DAYS_BACK
    status['traccar_cache']     = CACHE_DURATION

    return jsonify(status)


# Sauvegarde automatique de la base : extrait dans backup.py (importee en tete de
# fichier), voir 'from backup import backup_database, backup_scheduler'.




@app.route("/export_phyto_excel")
@login_required
def export_phyto_excel():
    """Export Excel registre phyto/semis — équivalent lisible du XML."""
    DB_PATH = 'database.db'
    TYPES_AUTORISES_REGISTRE = {'phyto', 'semence'}

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT device_id, geofence_id, exit_time,
                   intervention_type, products, applied_area, meteo
            FROM interventions
            WHERE intervention_type IN ('Pulvérisation', 'Semis')
            ORDER BY exit_time DESC
        """)
        interventions = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT name, type, amm, unit, culture, bbch, target, bio FROM catalog_products")
        catalog = {r['name']: dict(r) for r in cur.fetchall()}

        cur.execute("SELECT nom, code_oepp FROM cultures")
        cultures_oepp = {r['nom'].strip().lower(): r['code_oepp'] for r in cur.fetchall()}

        cur.execute("SELECT geofence_id, identifiant FROM parcelles")
        parcelles = {str(r['geofence_id']): r['identifiant'] for r in cur.fetchall()}

        cur.execute("SELECT siret, raison_sociale, applicateur, certiphyto FROM exploitation WHERE id = 1")
        row = cur.fetchone()
        siret = row['siret'] if row else ''
        raison_sociale = row['raison_sociale'] if row else ''
        applicateur = row['applicateur'] if row else ''
        certiphyto = row['certiphyto'] if row else ''

    raw = build_data()
    geofences = raw.get("geofences", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Registre Phyto"

    # En-tête exploitation
    ws.append(["SIRET", siret, "Raison sociale", raison_sociale])
    ws.append(["Applicateur", applicateur, "N° Certiphyto", certiphyto])
    ws.append([])
    ws.append(["Date", "Type", "ID Parcelle", "Nom Parcelle", "Surface (ha)",
               "Produit", "N° AMM", "Dose", "Cible", "Code culture", "Stade BBCH", "Bio"])

    for interv in interventions:
        geo_id_str = str(interv['geofence_id'])
        geo_name = geofences.get(geo_id_str, {}).get('name', f"Parcelle {interv['geofence_id']}")
        id_parcelle = parcelles.get(geo_id_str, '') or geo_name

        try:
            date_only = datetime.strptime(interv['exit_time'][:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            date_only = interv['exit_time'][:10]

        surf = interv.get('applied_area', '')

        try:
            products = json.loads(interv['products']) if interv['products'] else []
        except Exception:
            products = []

        if products:
            lignes_ecrites = 0
            for prod in products:
                prod_name = prod.get('name', '')
                cat = catalog.get(prod_name, {})
                # Le registre phyto exclut les engrais
                if cat.get('type') not in TYPES_AUTORISES_REGISTRE:
                    continue
                prod_dose = prod.get('dosage', '')
                unit = cat.get('unit', '')
                dose_display = f"{prod_dose} {unit}/ha".strip() if prod_dose else ''
                ws.append([
                    date_only, interv['intervention_type'], id_parcelle, geo_name, surf,
                    prod_name, cat.get('amm', ''), dose_display,
                    cat.get('target', ''), (cultures_oepp.get((cat.get('culture') or '').strip().lower(), '') or cat.get('culture', '')), cat.get('bbch', ''),
                    'Oui' if cat.get('bio') else 'Non'
                ])
                lignes_ecrites += 1
            if lignes_ecrites == 0:
                continue  # intervention 100% engrais : on ne l'inclut pas du tout
        else:
            continue  # pas de produits du tout : rien à mettre dans le registre phyto

    # Largeur des colonnes
    for col, width in zip('ABCDEFGHIJKL', [12,14,12,20,12,20,12,14,16,12,10,8]):
        ws.column_dimensions[col].width = width

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name="registre_phyto.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def get_cultures_rules():
    """Retourne {nom_culture: {debut_mmdd, fin_mmdd}}."""
    DB_PATH = 'database.db'
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT nom, debut_mmdd, fin_mmdd FROM cultures")
        return {r['nom'].strip().lower(): {'debut_mmdd': r['debut_mmdd'], 'fin_mmdd': r['fin_mmdd']} for r in cur.fetchall()}


def campagne_label_in_bounds(dt, debut_mmdd, fin_mmdd):
    """
    Calcule le label de campagne auquel appartient dt, selon les bornes MM-DD d'une culture.
    La campagne est TOUJOURS désignée par son année de récolte (ex: un blé semé en octobre
    2025 et récolté en juillet 2026 appartient à la "campagne 2026", pas "2025-2026") -- un
    label à cheval sur deux années aurait empêché toute comparaison directe avec l'année
    civile en cours ailleurs dans l'application (Synthèse, phénologie...).
    - Campagne calendaire simple (ex: maïs jan->déc) : retourne "2025" (année de récolte = dt.year)
    - Campagne hivernale à cheval (ex: blé sept N -> août N+1) : retourne "2026" (année de
      récolte, pas l'année de semis)
    """
    dmm, ddd = map(int, debut_mmdd.split('-'))
    fmm, fdd = map(int, fin_mmdd.split('-'))

    debut_avant_fin_meme_annee = (dmm, ddd) <= (fmm, fdd)

    if debut_avant_fin_meme_annee:
        # Campagne calendaire simple : label = l'année de dt (= année de récolte)
        return str(dt.year)
    else:
        # Campagne hivernale qui traverse le 1er janvier : label = année de RÉCOLTE
        # (année de début de campagne + 1), pas l'année de semis.
        seuil = datetime(dt.year, dmm, ddd)
        annee_debut = dt.year if dt >= seuil else dt.year - 1
        return str(annee_debut + 1)


def _resolve_culture_from_semis(semis_interv, products_catalog, cultures_rules, dt_semis, dt_interv):
    """Essaie de résoudre (culture, campagne) à partir des produits d'un semis candidat,
    en vérifiant que la date de l'intervention tombe dans les mêmes bornes de campagne."""
    try:
        products = json.loads(semis_interv.get('products') or '[]')
    except Exception:
        products = []
    for prod in products:
        prod_name = (prod.get('name') or '').strip()
        if not prod_name:
            continue
        cat_entry = products_catalog.get(prod_name, {})
        culture_nom = (cat_entry.get('culture') or '').strip()
        if not culture_nom:
            continue
        rule = cultures_rules.get(culture_nom.lower())
        if not rule:
            continue
        label_semis = campagne_label_in_bounds(dt_semis, rule['debut_mmdd'], rule['fin_mmdd'])
        label_interv = campagne_label_in_bounds(dt_interv, rule['debut_mmdd'], rule['fin_mmdd'])
        if label_semis == label_interv:
            return culture_nom, label_semis
    return None, None


def find_culture_for_intervention(geofence_id, exit_time_str, all_interventions, cultures_rules, products_catalog=None, sous_parcelle_id=None, sous_parcelles_info=None):
    """
    Détermine la culture d'une intervention en cherchant le semis dont elle dépend, sur la
    même parcelle, et en vérifiant que la date de l'intervention tombe dans les bornes de
    campagne de cette culture.
    La culture d'un semis est résolue via le CATALOGUE PRODUITS (par nom de produit),
    car le champ 'culture' n'est pas stocké directement dans le produit de l'intervention.

    Cas particulier PRIORITAIRE : si l'intervention est rattachée à une sous-parcelle
    (sous_parcelle_id), la culture et la campagne sont lues DIRECTEMENT depuis la
    sous-parcelle elle-même (déjà connues sans ambiguïté, saisies à sa création) plutôt que
    par proximité temporelle avec un semis. Sans ça, une parcelle scindée en plusieurs
    cultures (ex: maïs semé le 1er mai sur une zone, tournesol semé le 15 mai sur une autre)
    ferait retomber une intervention sur le semis le plus proche dans le temps, quelle que
    soit la sous-parcelle à laquelle elle appartient réellement -- mélangeant les deux
    cultures ou scindant à tort l'historique d'une même sous-parcelle en plusieurs groupes
    selon qu'une intervention tombe avant ou après tel ou tel semis "voisin".

    Sinon (pas de sous-parcelle), deux cas, dans cet ordre :
    1) Le DERNIER semis PASSÉ (ou le jour même) sur cette parcelle -- c'est le cas normal pour
       toute intervention qui suit son semis (épandage, pulvérisation, récolte...). Sans ce
       cas, une intervention réalisée après son propre semis ne pouvait jamais s'y rattacher,
       et retombait dans une autre campagne que celle de sa culture réelle.
    2) À défaut, le PROCHAIN semis à venir -- cas des opérations de préparation réalisées
       avant la plantation (labour, hersage...), qui doivent se rattacher à la campagne à venir.
    Dans ce cas de repli, seuls les semis NON rattachés à une sous-parcelle sont considérés
    (semis concernant la parcelle entière) : le semis d'une sous-parcelle particulière ne
    doit pas "happer" une intervention qui, elle, ne lui est pas rattachée.
    Retourne (nom_culture, campagne_label) ou (None, None) si indéterminé.
    """
    if products_catalog is None:
        products_catalog = {}

    if sous_parcelle_id and sous_parcelles_info:
        info = sous_parcelles_info.get(sous_parcelle_id)
        if info and info.get("culture"):
            return info["culture"], (info.get("campagne") or None)

    try:
        dt_interv = datetime.strptime(exit_time_str[:19], "%Y-%m-%dT%H:%M:%S")
    except Exception:
        try:
            dt_interv = datetime.strptime(exit_time_str[:10], "%Y-%m-%d")
        except Exception:
            return None, None

    semis_passes = []
    semis_futurs = []
    for other in all_interventions:
        if other.get('geofence_id') != geofence_id:
            continue
        if other.get('intervention_type') != 'Semis':
            continue
        if other.get('sous_parcelle_id'):
            # Semis d'une sous-parcelle particulière : ne concerne pas une intervention qui
            # n'y est elle-même pas rattachée (sous_parcelle_id est None ici, sinon on
            # serait déjà sortis via le cas prioritaire ci-dessus).
            continue
        try:
            dt_semis = datetime.strptime(other['exit_time'][:19], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            try:
                dt_semis = datetime.strptime(other['exit_time'][:10], "%Y-%m-%d")
            except Exception:
                continue
        if dt_semis <= dt_interv:
            semis_passes.append((dt_semis, other))
        else:
            semis_futurs.append((dt_semis, other))

    # 1) Semis passés : du plus récent au plus ancien (le plus proche a la priorité)
    semis_passes.sort(key=lambda x: x[0], reverse=True)
    for dt_semis, semis_interv in semis_passes:
        culture_nom, label = _resolve_culture_from_semis(semis_interv, products_catalog, cultures_rules, dt_semis, dt_interv)
        if culture_nom:
            return culture_nom, label

    # 2) Repli : semis à venir, du plus proche au plus lointain
    semis_futurs.sort(key=lambda x: x[0])
    for dt_semis, semis_interv in semis_futurs:
        culture_nom, label = _resolve_culture_from_semis(semis_interv, products_catalog, cultures_rules, dt_semis, dt_interv)
        if culture_nom:
            return culture_nom, label

    return None, None


@app.route("/api/interventions_campagnes")
@login_required
def api_interventions_campagnes():
    """
    Pour chaque intervention, résout sa campagne (culture + label de campagne) via la même
    logique que le Bilan et la Fertilisation (find_culture_for_intervention), afin que le
    Carnet puisse filtrer par campagne sans dupliquer/désynchroniser cette résolution.
    Renvoie {clé "device_id_geofence_id_exit_time": {"culture":..., "campagne":...}}.
    """
    DB_PATH = 'database.db'
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT device_id, geofence_id, exit_time, intervention_type, products FROM interventions")
        all_interventions = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT name, culture FROM catalog_products")
        products_catalog = {r['name'].strip(): dict(r) for r in cur.fetchall()}

    cultures_rules = get_cultures_rules()
    result = {}
    for interv in all_interventions:
        culture_nom, campagne_label = find_culture_for_intervention(
            interv['geofence_id'], interv['exit_time'], all_interventions, cultures_rules, products_catalog
        )
        key = f"{interv['device_id']}_{interv['geofence_id']}_{interv['exit_time']}"
        result[key] = {"culture": culture_nom, "campagne": campagne_label}
    return jsonify(result)





# =========================================================================
# ROUTE CONFIG TRACCAR
# =========================================================================

@app.route("/api/config_traccar", methods=["GET", "POST"])
@login_required
def api_config_traccar():
    global TRACCAR_URL, TRACCAR_USER, TRACCAR_PASSWORD, DAYS_BACK, CACHE_DURATION, _cached_data, _last_cache_time
    import json as _json
    config_path = "config.json"
    if request.method == "GET":
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        tcfg = cfg.get("traccar", {})
        return jsonify({"url": tcfg.get("url", TRACCAR_URL), "user": tcfg.get("user", TRACCAR_USER),
                        "password": "", "days_back": tcfg.get("days_back", DAYS_BACK),
                        "cache_duration": tcfg.get("cache_duration", CACHE_DURATION)})
    else:
        data = request.get_json()
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        cfg["traccar"] = {"url": data.get("url","").strip(), "user": data.get("user","").strip(),
                          "password": data.get("password","").strip(),
                          "days_back": int(data.get("days_back", 30)),
                          "cache_duration": int(data.get("cache_duration", 60))}
        with open(config_path, "w") as f: _json.dump(cfg, f, indent=2)
        TRACCAR_URL = cfg["traccar"]["url"]; TRACCAR_USER = cfg["traccar"]["user"]
        if cfg["traccar"]["password"]: TRACCAR_PASSWORD = cfg["traccar"]["password"]
        DAYS_BACK = cfg["traccar"]["days_back"]; CACHE_DURATION = cfg["traccar"]["cache_duration"]
        session.auth = (TRACCAR_USER, TRACCAR_PASSWORD)
        _cached_data = None; _last_cache_time = 0
        return jsonify({"status": "ok", "message": "Configuration Traccar mise à jour et appliquée."})


@app.route("/api/config_partage", methods=["GET", "POST"])
@login_required
def api_config_partage():
    """
    Modele de message (prefixe/suffixe) ajoute automatiquement autour du contenu genere lors
    d'un partage par email/SMS/WhatsApp (points d'observation, parcelles, fiches
    d'intervention...). Suit le meme pattern de stockage que /api/config_traccar (config.json),
    par coherence avec le reste des reglages de l'application plutot que d'introduire un
    nouveau mecanisme de stockage.
    """
    import json as _json
    config_path = "config.json"
    if request.method == "GET":
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        pcfg = cfg.get("partage", {})
        return jsonify({"prefix": pcfg.get("prefix", ""), "suffix": pcfg.get("suffix", "")})
    else:
        data = request.get_json(silent=True) or {}
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        cfg["partage"] = {"prefix": str(data.get("prefix", "")).strip(),
                           "suffix": str(data.get("suffix", "")).strip()}
        with open(config_path, "w") as f: _json.dump(cfg, f, indent=2)
        return jsonify({"status": "ok"})


# =========================================================================
# ALERTES EMAIL AUTOMATIQUES (SMTP)
# =========================================================================
# Contrairement au partage manuel (mailto:/sms:/WhatsApp, qui necessite un clic humain),
# ce mecanisme permet au SERVEUR d'envoyer lui-meme un email sans aucune action de
# l'utilisateur -- necessaire pour de vraies alertes automatiques (ex: Traccar tombe en
# panne a 3h du matin, personne n'a le dashboard ouvert dans un navigateur a ce moment-la).

def _charger_config_smtp():
    import json as _json
    cfg = {}
    try:
        if os.path.exists("config.json"):
            with open("config.json", "r") as f: cfg = _json.load(f)
    except Exception: pass
    return cfg.get("smtp", {})


def envoyer_email(destinataire, sujet, corps):
    """
    Envoie un email via le serveur SMTP configure (config.json, section "smtp"). Retourne
    (True, "") en cas de succes, (False, message_erreur) sinon -- ne leve jamais
    d'exception : un envoi d'alerte qui echoue (mauvais mot de passe, serveur SMTP
    injoignable...) ne doit pas faire planter le reste de l'application.
    """
    smtp_cfg = _charger_config_smtp()
    host = smtp_cfg.get("host")
    port = smtp_cfg.get("port", 587)
    user = smtp_cfg.get("user")
    password = smtp_cfg.get("password")
    expediteur = smtp_cfg.get("from") or user

    if not host or not user or not password:
        return False, "Configuration SMTP incomplète (voir ⚙️ État système)"
    if not destinataire:
        return False, "Aucun destinataire configuré pour les alertes"

    msg = MIMEMultipart()
    msg["From"] = expediteur
    msg["To"] = destinataire
    msg["Subject"] = sujet
    msg.attach(MIMEText(corps, "plain", "utf-8"))

    try:
        with smtplib.SMTP(host, int(port), timeout=15) as server:
            server.starttls()
            server.login(user, password)
            server.sendmail(expediteur, [destinataire], msg.as_string())
        return True, ""
    except Exception as e:
        return False, str(e)


@app.route("/api/config_smtp", methods=["GET", "POST"])
@login_required
def api_config_smtp():
    """
    Configuration du serveur SMTP et des alertes automatiques activees. Le mot de passe
    n'est jamais renvoye en clair au GET (meme principe que /api/config_traccar) -- au
    POST, un mot de passe vide conserve l'ancien plutot que de l'effacer.
    """
    import json as _json
    config_path = "config.json"
    if request.method == "GET":
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        scfg = cfg.get("smtp", {})
        return jsonify({
            "host": scfg.get("host", ""), "port": scfg.get("port", 587),
            "user": scfg.get("user", ""), "password": "",
            "from": scfg.get("from", ""),
            "destinataire_alertes": scfg.get("destinataire_alertes", ""),
            "alerte_traccar": scfg.get("alerte_traccar", False),
            "alerte_points_terrain": scfg.get("alerte_points_terrain", False),
        })
    else:
        data = request.get_json(silent=True) or {}
        cfg = {}
        try:
            if os.path.exists(config_path):
                with open(config_path, "r") as f: cfg = _json.load(f)
        except Exception: pass
        ancien_smtp = cfg.get("smtp", {})
        nouveau_mdp = str(data.get("password", "")).strip()
        cfg["smtp"] = {
            "host": str(data.get("host", "")).strip(),
            "port": int(data.get("port", 587) or 587),
            "user": str(data.get("user", "")).strip(),
            "password": nouveau_mdp if nouveau_mdp else ancien_smtp.get("password", ""),
            "from": str(data.get("from", "")).strip(),
            "destinataire_alertes": str(data.get("destinataire_alertes", "")).strip(),
            "alerte_traccar": bool(data.get("alerte_traccar", False)),
            "alerte_points_terrain": bool(data.get("alerte_points_terrain", False)),
        }
        with open(config_path, "w") as f: _json.dump(cfg, f, indent=2)
        return jsonify({"status": "ok"})


@app.route("/api/test_email", methods=["POST"])
@login_required
def api_test_email():
    """Envoie un email de test au destinataire configuré, pour valider la config SMTP."""
    smtp_cfg = _charger_config_smtp()
    destinataire = smtp_cfg.get("destinataire_alertes")
    ok, err = envoyer_email(
        destinataire, "Test — Dashboard Agricole",
        "Ceci est un email de test envoyé depuis le Dashboard Agricole.\n\n"
        "Si vous le recevez, la configuration SMTP fonctionne correctement."
    )
    if ok:
        return jsonify({"status": "ok"})
    return jsonify({"error": err}), 400


# Etat precedent de Traccar (accessible ou non), pour ne notifier QUE lors d'un changement
# d'etat -- sans cette memoire, une verification periodique enverrait un email a chaque
# cycle tant que Traccar reste en panne (spam), plutot qu'une seule fois a la bascule.
_traccar_etait_accessible = True

def traccar_health_alert_scheduler():
    """
    Verifie periodiquement l'accessibilite de Traccar en arriere-plan, independamment de
    toute requete HTTP d'un navigateur (contrairement a /api/system_status, qui ne verifie
    que lorsqu'un client interroge activement cette route) -- necessaire pour etre alerte
    meme si personne n'a le dashboard ouvert au moment de la panne.
    """
    global _traccar_etait_accessible
    while True:
        time.sleep(300)  # verification toutes les 5 minutes
        try:
            smtp_cfg = _charger_config_smtp()
            if not smtp_cfg.get("alerte_traccar"):
                continue

            try:
                test = safe_get(f"{TRACCAR_URL}/devices")
                accessible = bool(test)
            except Exception:
                accessible = False

            if accessible != _traccar_etait_accessible:
                destinataire = smtp_cfg.get("destinataire_alertes")
                if not accessible:
                    envoyer_email(destinataire, "⚠️ Traccar inaccessible",
                                  f"Le serveur Traccar ({TRACCAR_URL}) ne répond plus depuis le "
                                  f"Dashboard Agricole. Vérifiez la connexion réseau et les "
                                  f"identifiants.")
                else:
                    envoyer_email(destinataire, "✅ Traccar de nouveau accessible",
                                  f"Le serveur Traccar ({TRACCAR_URL}) répond de nouveau "
                                  f"normalement.")
                _traccar_etait_accessible = accessible
        except Exception:
            pass  # ne jamais laisser planter ce thread de fond


threading.Thread(target=traccar_health_alert_scheduler, daemon=True).start()


# =========================================================================
# EXPORT PDF — CAHIER DE FERTILISATION
# =========================================================================



# =========================================================================
# EXPORT PDF — ANALYTIQUE
# =========================================================================




# e-phy (ANSES) : synchronisation et recherche produits phytosanitaires -- extrait dans
# ephy_bp.py, voir l'enregistrement du blueprint en tete de fichier.


# =========================================================================
# CORRIDORS — Carte de travail à largeur réelle (RTK)
# =========================================================================

@app.route("/api/devices")
@login_required
def api_devices():
    """Retourne la liste des véhicules Traccar avec id et nom."""
    devices = get_devices()
    return jsonify([
        {"id": str(d_id), "name": info.get("name", f"Device {d_id}")}
        for d_id, info in devices.items()
    ])


def _parse_geofence_wkt(area):
    """
    Parse le champ 'area' (WKT) d'une géofence Traccar.
    Formats gérés : POLYGON ((lat lon, lat lon, ...)), CIRCLE (lat lon, radius_m).
    NB: Traccar stocke ici l'ordre (lat, lon) et non l'ordre OGC standard (lon, lat).
    Retourne {"type": "polygon", "coords": [[lat,lon], ...]} ou
             {"type": "circle", "center": [lat,lon], "radius": m} ou None.
    """
    if not area or not isinstance(area, str):
        return None
    s = area.strip()
    try:
        if s.upper().startswith("POLYGON"):
            inner = s[s.find("((")+2 : s.rfind("))")]
            coords = []
            for pair in inner.split(","):
                parts = pair.strip().split()
                if len(parts) >= 2:
                    lat, lon = float(parts[0]), float(parts[1])
                    coords.append([lat, lon])
            return {"type": "polygon", "coords": coords} if len(coords) >= 3 else None
        elif s.upper().startswith("CIRCLE"):
            inner = s[s.find("(")+1 : s.rfind(")")]
            center_part, radius_part = inner.rsplit(",", 1)
            lat, lon = [float(x) for x in center_part.strip().split()]
            radius = float(radius_part.strip())
            return {"type": "circle", "center": [lat, lon], "radius": radius}
    except Exception:
        return None
    return None


@app.route("/api/geofences")
@login_required
def api_geofences():
    """Retourne les géofences (parcelles) Traccar avec leur géométrie pour affichage carte."""
    data = safe_get(f"{TRACCAR_URL}/geofences")
    if not isinstance(data, list):
        return jsonify([])

    result = []
    for g in data:
        geom = _parse_geofence_wkt(g.get("area"))
        if not geom:
            continue
        result.append({
            "id": g.get("id"),
            "name": g.get("name", f"Parcelle {g.get('id')}"),
            **geom
        })
    return jsonify(result)


# ================= ANALYSE DE COUVERTURE (parcelle vs passages travaillés) =================
def _haversine_m(lat1, lon1, lat2, lon2):
    import math
    R = 6371000
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def _point_in_polygon(px, py, poly):
    """Ray casting standard. poly = liste de [x, y] (ici [lat, lon], peu importe l'ordre pourvu qu'il soit cohérent)."""
    n = len(poly)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if ((yi > py) != (yj > py)):
            x_intersect = (xj - xi) * (py - yi) / ((yj - yi) or 1e-15) + xi
            if px < x_intersect:
                inside = not inside
        j = i
    return inside

def _bbox_of(coords):
    lats = [c[0] for c in coords]; lons = [c[1] for c in coords]
    return min(lats), max(lats), min(lons), max(lons)

@app.route("/api/coverage")
@login_required
def api_coverage():
    """
    Analyse de couverture : compare la surface d'une parcelle (géofence) à ce qui a été
    effectivement travaillé (corridors) sur une date -- ou une période -- donnée, pour un
    ou plusieurs véhicules.
    Paramètres : geofence_id, device_ids (séparés par des virgules), resolution_m (def. 8), et :
      - soit "date" (YYYY-MM-DD) pour une seule journée (comportement historique)
      - soit "start_date" + "end_date" pour cumuler la couverture sur toute une période
        (ex: vérifier qu'une parcelle a été intégralement couverte sur toute la campagne)
    """
    geofence_id = request.args.get("geofence_id")
    device_ids  = [d for d in request.args.get("device_ids", "").split(",") if d]
    resolution_m = max(2.0, _to_float(request.args.get("resolution_m", 8), 8))
    sous_parcelle_id = request.args.get("sous_parcelle_id")
    sous_parcelle_id = int(sous_parcelle_id) if sous_parcelle_id and sous_parcelle_id.isdigit() else None

    start_date = request.args.get("start_date")
    end_date   = request.args.get("end_date")
    if start_date and end_date:
        try:
            dates = _daterange_str(start_date, end_date)
        except ValueError:
            return jsonify({"error": "Dates invalides, format YYYY-MM-DD"}), 400
    else:
        dates = [request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))]

    if not geofence_id or not device_ids:
        return jsonify({"error": "geofence_id et device_ids requis"}), 400

    MAX_JOBS = 200  # garde-fou : evite une periode demesurement longue x trop de vehicules
    if len(dates) * len(device_ids) > MAX_JOBS:
        return jsonify({"error": f"Trop de combinaisons véhicule × jour ({len(dates)*len(device_ids)} > {MAX_JOBS}). Réduisez la période ou le nombre de véhicules."}), 400

    result, status = _compute_coverage(geofence_id, dates, device_ids, resolution_m, sous_parcelle_id=sous_parcelle_id)
    return jsonify(result), status


def _compute_coverage(geofence_id, dates, device_ids, resolution_m=8.0, sous_parcelle_id=None):
    """
    Analyse de couverture : compare la surface d'une parcelle (géofence) -- ou d'une
    sous-parcelle si sous_parcelle_id est fourni -- à ce qui a été effectivement travaillé
    (corridors) sur une date donnée, pour un ou plusieurs véhicules.
    Méthode : échantillonnage par grille régulière (pas de dépendance externe type Shapely) --
    suffisamment précis pour du pilotage terrain à une résolution de quelques mètres.
    Retourne (dict_resultat, code_http).
    """
    import math

    if isinstance(dates, str):
        dates = [dates]

    # ── Récupérer la géométrie : celle de la sous-parcelle si précisée, sinon la parcelle entière ──
    if sous_parcelle_id:
        with sqlite3.connect('database.db') as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT polygon FROM sous_parcelles WHERE id = ?", (sous_parcelle_id,)).fetchone()
        if not row:
            return {"error": "Sous-parcelle introuvable"}, 404
        try:
            coords = json.loads(row["polygon"])
        except Exception:
            coords = []
        if len(coords) < 3:
            return {"error": "Polygone de sous-parcelle invalide"}, 422
        geom = {"type": "polygon", "coords": coords}
    else:
        raw_geofences = safe_get(f"{TRACCAR_URL}/geofences")
        if not isinstance(raw_geofences, list):
            return {"error": "Impossible de récupérer les géofences"}, 502
        gf = next((g for g in raw_geofences if str(g.get("id")) == str(geofence_id)), None)
        if not gf:
            return {"error": "Géofence introuvable"}, 404
        geom = _parse_geofence_wkt(gf.get("area"))
        if not geom:
            return {"error": "Géométrie de la géofence illisible"}, 422

    if geom["type"] == "polygon":
        poly = geom["coords"]
        lat_min, lat_max, lon_min, lon_max = _bbox_of(poly)
    else:  # circle
        clat, clon = geom["center"]; radius = geom["radius"]
        deg_lat = radius / 111320.0
        deg_lon = radius / (111320.0 * max(0.0001, math.cos(math.radians(clat))))
        lat_min, lat_max = clat - deg_lat, clat + deg_lat
        lon_min, lon_max = clon - deg_lon, clon + deg_lon

    def _inside_geofence(lat, lon):
        if geom["type"] == "polygon":
            return _point_in_polygon(lat, lon, poly)
        return _haversine_m(lat, lon, geom["center"][0], geom["center"][1]) <= geom["radius"]

    # ── Récupérer tous les quadrilatères travaillés (tous véhicules et toutes dates sélectionnés) ──
    all_quads = []   # [(bbox, quad_coords), ...]
    all_joints = []  # [(lat, lon, radius_m), ...]
    per_device_area = {}
    for dev_id in device_ids:
        for d in dates:
            pts = _fetch_and_parse_positions(dev_id, d)
            if not pts:
                continue
            try:
                data = _build_corridors_response(pts, d, math)
            except Exception:
                continue
            per_device_area[dev_id] = per_device_area.get(dev_id, 0) + data.get("stats", {}).get("area_ha", 0)
            for corridor in data.get("corridors", []):
                for q in corridor.get("quads", []):
                    all_quads.append((_bbox_of(q), q))
                radius = corridor.get("joint_radius_m", 0)
                for j in corridor.get("joints", []):
                    all_joints.append((j[0], j[1], radius))

    if not all_quads and not all_joints:
        return {
            "coverage_pct": 0, "geofence_area_ha": None, "covered_area_ha": 0,
            "uncovered_points": [], "truncated": False,
            "message": "Aucun passage travaillé trouvé pour cette sélection (véhicules/période)."
        }, 200

    # ── Grille régulière sur la bbox de la parcelle ──
    deg_step_lat = resolution_m / 111320.0
    mid_lat = (lat_min + lat_max) / 2.0
    deg_step_lon = resolution_m / (111320.0 * max(0.0001, math.cos(math.radians(mid_lat))))

    nb_lat = max(1, int((lat_max - lat_min) / deg_step_lat) + 1)
    nb_lon = max(1, int((lon_max - lon_min) / deg_step_lon) + 1)
    # Garde-fou : limite le nombre total de points echantillonnes (grande parcelle / faible resolution)
    MAX_GRID_POINTS = 40000
    if nb_lat * nb_lon > MAX_GRID_POINTS:
        scale = math.sqrt((nb_lat * nb_lon) / MAX_GRID_POINTS)
        deg_step_lat *= scale
        deg_step_lon *= scale
        nb_lat = max(1, int((lat_max - lat_min) / deg_step_lat) + 1)
        nb_lon = max(1, int((lon_max - lon_min) / deg_step_lon) + 1)

    total_in = 0
    covered_in = 0
    uncovered_points = []
    MAX_RETURNED_POINTS = 3000

    for ilat in range(nb_lat + 1):
        lat = lat_min + ilat * deg_step_lat
        for ilon in range(nb_lon + 1):
            lon = lon_min + ilon * deg_step_lon
            if not _inside_geofence(lat, lon):
                continue
            total_in += 1

            covered = False
            for (bbox, quad) in all_quads:
                bl_min, bl_max, bo_min, bo_max = bbox
                if lat < bl_min or lat > bl_max or lon < bo_min or lon > bo_max:
                    continue
                if _point_in_polygon(lat, lon, quad):
                    covered = True
                    break
            if not covered:
                for (jlat, jlon, jradius) in all_joints:
                    if _haversine_m(lat, lon, jlat, jlon) <= jradius:
                        covered = True
                        break

            if covered:
                covered_in += 1
            elif len(uncovered_points) < MAX_RETURNED_POINTS:
                uncovered_points.append([round(lat, 6), round(lon, 6)])

    coverage_pct = round(100.0 * covered_in / total_in, 1) if total_in else 0
    cell_area_ha = (resolution_m * resolution_m) / 10000.0
    geofence_area_ha = round(total_in * cell_area_ha, 2)
    covered_area_ha = round(covered_in * cell_area_ha, 2)

    return {
        "coverage_pct": coverage_pct,
        "geofence_area_ha": geofence_area_ha,
        "covered_area_ha": covered_area_ha,
        "uncovered_area_ha": round(geofence_area_ha - covered_area_ha, 2),
        "resolution_m": resolution_m,
        "uncovered_points": uncovered_points,
        "truncated": total_in - covered_in > MAX_RETURNED_POINTS,
        "per_device_area_ha": per_device_area,
        "alert_threshold_pct": ALERT_MIN_COVERAGE_PCT,
    }, 200


@app.route("/api/chantiers_calendar")
@login_required
def api_chantiers_calendar():
    """
    Retourne pour chaque date les véhicules ayant travaillé
    (isWorking=True dans au moins une position).
    Paramètres : month (YYYY-MM)
    """
    month = request.args.get("month", datetime.utcnow().strftime("%Y-%m"))
    try:
        dt_month = datetime.strptime(month, "%Y-%m")
    except ValueError:
        return jsonify({"error": "format YYYY-MM requis"}), 400

    # Récupérer sur tout le mois
    start_str = dt_month.strftime("%Y-%m-%dT00:00:00Z")
    import calendar as _cal
    last_day = _cal.monthrange(dt_month.year, dt_month.month)[1]
    end_str = dt_month.strftime(f"%Y-%m-{last_day:02d}T23:59:59Z")

    devices = get_devices()
    result = {}  # date -> list of {id, name}

    import concurrent.futures as _cf

    def fetch_device(d_id, d_info):
        url = f"{TRACCAR_URL}/reports/route?deviceId={d_id}&from={start_str}&to={end_str}"
        positions = safe_get(url)
        if not isinstance(positions, list): return {}
        days = {}
        for p in positions:
            attrs = p.get("attributes", {})
            if not to_bool(attrs.get("isWorking", False)): continue
            fix_time = p.get("fixTime", "")
            if not fix_time: continue
            try:
                day = fix_time[:10]  # YYYY-MM-DD
                if day not in days:
                    days[day] = {"id": str(d_id), "name": d_info.get("name", f"Device {d_id}")}
            except Exception: pass
        return days

    with _cf.ThreadPoolExecutor(max_workers=6) as ex:
        futures = {ex.submit(fetch_device, d_id, d_info): d_id
                   for d_id, d_info in devices.items()}
        for fut in _cf.as_completed(futures):
            days = fut.result()
            for day, info in days.items():
                if day not in result: result[day] = []
                # Eviter les doublons
                if not any(v["id"] == info["id"] for v in result[day]):
                    result[day].append(info)

    return jsonify(result)


@app.route("/api/corridors")
@login_required
def api_corridors():
    """
    Récupère les positions Traccar d'un device sur une journée
    et calcule les corridors de travail à largeur réelle.
    Paramètres : device_id, date (YYYY-MM-DD)
    """
    import math

    device_id = request.args.get("device_id")
    date_str   = request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))

    if not device_id:
        return jsonify({"error": "device_id requis"}), 400

    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "date invalide, format YYYY-MM-DD"}), 400

    pts = _fetch_and_parse_positions(device_id, date_str)
    if not pts:
        return jsonify({"corridors": [], "stats": {}, "positions": []})

    try:
        return jsonify(_build_corridors_response(pts, date_str, math))
    except Exception as e:
        app.logger.exception("api_corridors: erreur calcul corridors")
        return jsonify({"error": f"Erreur calcul corridors: {e}"}), 500


def _to_float(val, default=0.0):
    """Convertit en float, tolère virgule décimale et unité collée (ex: '4,5 m', '3.2m')."""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        pass
    s = str(val).strip().lower().replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if m:
        try:
            return float(m.group())
        except ValueError:
            return default
    return default


def _fetch_and_parse_positions(device_id, date_str):
    """Récupère et normalise les positions Traccar d'un device pour une date donnée (YYYY-MM-DD)."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    start_str = dt.strftime("%Y-%m-%dT00:00:00Z")
    end_str   = dt.strftime("%Y-%m-%dT23:59:59Z")

    url = f"{TRACCAR_URL}/reports/route?deviceId={device_id}&from={start_str}&to={end_str}"
    positions = safe_get(url)
    if not isinstance(positions, list) or len(positions) == 0:
        return []

    pts = []
    for p in positions:
        lat  = p.get("latitude")
        lon  = p.get("longitude")
        spd  = p.get("speed", 0) or 0   # km/h dans Traccar
        attrs = p.get("attributes", {})
        is_working    = to_bool(attrs.get("isWorking", False))
        working_width = _to_float(attrs.get("workingWidth") or attrs.get("width") or 0)
        tool          = str(attrs.get("tool", "") or "").strip()
        fix_time = p.get("fixTime", "")
        if lat is None or lon is None: continue
        pts.append({
            "lat": _to_float(lat), "lon": _to_float(lon),
            "speed": _to_float(spd), "isWorking": is_working,
            "width": working_width, "tool": tool, "time": fix_time
        })
    return pts


def _build_corridors_response(pts, date_str, math):
    # ── Construire les corridors (segments de travail actif) ──
    def latlon_to_meters(lat1, lon1, lat2, lon2):
        """Distance approx en mètres entre deux points GPS."""
        R = 6371000
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

    def offset_point(lat, lon, bearing_deg, dist_m):
        """Déplace un point de dist_m dans la direction bearing_deg."""
        R = 6371000
        d = dist_m / R
        b = math.radians(bearing_deg)
        lat1 = math.radians(lat)
        lon1 = math.radians(lon)
        _asin_arg = math.sin(lat1)*math.cos(d) + math.cos(lat1)*math.sin(d)*math.cos(b)
        _asin_arg = max(-1.0, min(1.0, _asin_arg))  # évite ValueError: math domain error (arrondis flottants)
        lat2 = math.asin(_asin_arg)
        lon2 = lon1 + math.atan2(math.sin(b)*math.sin(d)*math.cos(lat1), math.cos(d)-math.sin(lat1)*math.sin(lat2))
        return math.degrees(lat2), math.degrees(lon2)

    def bearing(lat1, lon1, lat2, lon2):
        """Cap entre deux points en degrés."""
        dlon = math.radians(lon2 - lon1)
        y = math.sin(dlon) * math.cos(math.radians(lat2))
        x = math.cos(math.radians(lat1)) * math.sin(math.radians(lat2)) -             math.sin(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.cos(dlon)
        return (math.degrees(math.atan2(y, x)) + 360) % 360

    corridors  = []
    total_dist = 0.0
    total_area = 0.0
    work_times = []
    speeds_work = []

    i = 0
    while i < len(pts) - 1:
        p = pts[i]
        if not p["isWorking"]:
            i += 1
            continue

        # Début d'un segment de travail
        seg = [p]
        j = i + 1
        while j < len(pts) and pts[j]["isWorking"]:
            seg.append(pts[j])
            j += 1

        if len(seg) >= 2:
            # Largeur : utiliser la largeur du premier point valide du segment
            w = next((s["width"] for s in seg if s["width"] > 0), 3.0)
            half = w / 2.0

            # Outil : utilise le mot-clé outil (attribut Traccar "tool") du premier point valide
            # du segment -- transmis au frontend pour déduire le nom de l'intervention via le
            # catalogue outil -> intervention (même logique que la page principale index.html).
            tool = next((s["tool"] for s in seg if s.get("tool")), "")

            # Construire le polygone du corridor
            left_side  = []
            right_side = []

            for k in range(len(seg)):
                if k < len(seg) - 1:
                    b = bearing(seg[k]["lat"], seg[k]["lon"], seg[k+1]["lat"], seg[k+1]["lon"])
                else:
                    b = bearing(seg[k-1]["lat"], seg[k-1]["lon"], seg[k]["lat"], seg[k]["lon"])

                perp_left  = (b - 90) % 360
                perp_right = (b + 90) % 360

                ll = offset_point(seg[k]["lat"], seg[k]["lon"], perp_left,  half)
                lr = offset_point(seg[k]["lat"], seg[k]["lon"], perp_right, half)
                left_side.append(ll)
                right_side.append(lr)

            # Polygone : gauche aller + droite retour
            poly_coords = left_side + list(reversed(right_side))

            # Découpage en quadrilatères simples (un par intervalle de points GPS).
            # IMPORTANT : contrairement au polygone d'ensemble ci-dessus (qui utilise un cap
            # "par point", moyenné/anticipé pour lisser le contour global), chaque quadrilatère
            # est ici recalculé avec un SEUL cap partagé par ses deux extrémités (le cap du
            # segment k -> k+1). C'est ce qui garantit que le quadrilatère reste un simple
            # rectangle, jamais vrillé : si on réutilisait les caps "par point" (qui diffèrent
            # d'un point à l'autre dans un virage serré), le quadrilatère pouvait devenir
            # auto-sécant ("bowtie"), et le rendu SVG laissait alors la zone de croisement de
            # l'outil non coloriée. Avec un cap unique par segment, ce problème disparaît
            # totalement, y compris dans les virages serrés et les recroisements de l'outil.
            quads = []
            for k in range(len(seg) - 1):
                b_edge = bearing(seg[k]["lat"], seg[k]["lon"], seg[k+1]["lat"], seg[k+1]["lon"])
                perp_l = (b_edge - 90) % 360
                perp_r = (b_edge + 90) % 360
                l_k  = offset_point(seg[k]["lat"],   seg[k]["lon"],   perp_l, half)
                l_k1 = offset_point(seg[k+1]["lat"], seg[k+1]["lon"], perp_l, half)
                r_k  = offset_point(seg[k]["lat"],   seg[k]["lon"],   perp_r, half)
                r_k1 = offset_point(seg[k+1]["lat"], seg[k+1]["lon"], perp_r, half)
                quads.append([l_k, l_k1, r_k1, r_k])

            # Jointures : à chaque point intermédiaire (ni le tout premier, ni le tout dernier
            # du passage), les deux quadrilatères voisins ont chacun leur propre cap (celui de
            # leur propre segment). Dès que le tracteur tourne, ces deux caps diffèrent : côté
            # intérieur du virage les quadrilatères se chevauchent (pas de souci), mais côté
            # extérieur ils laissent un petit vide triangulaire entre eux (l'équivalent d'une
            # jointure "en biseau" non couverte). On comble ce vide avec un petit disque plein
            # centré sur le point GPS, de rayon = demi-largeur de l'outil : ça referme
            # complètement le ruban à chaque angle, quelle que soit la sévérité du virage.
            joints = [[seg[k]["lat"], seg[k]["lon"]] for k in range(1, len(seg) - 1)]

            # Distance et surface du segment
            seg_dist = sum(latlon_to_meters(seg[k]["lat"], seg[k]["lon"],
                                             seg[k+1]["lat"], seg[k+1]["lon"])
                           for k in range(len(seg)-1))
            seg_area = seg_dist * w / 10000  # ha

            total_dist += seg_dist
            total_area += seg_area

            spd_vals = [s["speed"] for s in seg if s["speed"] > 0]
            if spd_vals:
                speeds_work.extend(spd_vals)

            corridors.append({
                "polygon": poly_coords,  # [[lat,lon], ...] contour global (tooltip / KML / bounds)
                "quads":   quads,        # liste de petits quadrilatères simples pour l'affichage
                "joints":  joints,       # points d'articulation à combler (voir commentaire ci-dessus)
                "joint_radius_m": half,  # rayon des disques de jointure (= demi-largeur outil)
                "width":   w,
                "tool":    tool,         # mot-clé outil détecté (pour déduire le nom de l'intervention côté client)
                "dist_m":  round(seg_dist, 1),
                "area_ha": round(seg_area, 4),
                "active":  True
            })

        i = j

    avg_speed = round(sum(speeds_work) / len(speeds_work), 1) if speeds_work else 0

    # Positions brutes pour affichage trace GPS
    pos_out = [{"lat": p["lat"], "lon": p["lon"],
                "isWorking": p["isWorking"], "time": p["time"],
                "speed": round(p["speed"], 1)} for p in pts]

    # ── Alertes vitesse anormale : points travaillés (isWorking) au-delà du seuil configuré ──
    # (une vitesse excessive pendant le travail signale souvent une mauvaise application/dose)
    speed_alerts = [
        {"lat": p["lat"], "lon": p["lon"], "speed": round(p["speed"], 1), "time": p["time"]}
        for p in pts if p["isWorking"] and p["speed"] > ALERT_MAX_WORKING_SPEED_KMH
    ]

    # ── Alertes arrêt prolongé, outil engagé : le véhicule reste quasi immobile
    # (vitesse ~0) alors que l'outil est toujours actif -- signe possible d'un bourrage,
    # d'une panne, ou d'un oubli de relever l'outil pendant une pause.
    STOP_SPEED_THRESHOLD_KMH = 1.0
    stop_alerts = []
    i = 0
    while i < len(pts):
        p = pts[i]
        if p["isWorking"] and p["speed"] <= STOP_SPEED_THRESHOLD_KMH:
            j = i
            while j < len(pts) and pts[j]["isWorking"] and pts[j]["speed"] <= STOP_SPEED_THRESHOLD_KMH:
                j += 1
            try:
                t_start = datetime.strptime(pts[i]["time"][:19], "%Y-%m-%dT%H:%M:%S")
                t_end = datetime.strptime(pts[j-1]["time"][:19], "%Y-%m-%dT%H:%M:%S")
                duree_min = (t_end - t_start).total_seconds() / 60.0
            except Exception:
                duree_min = 0
            if duree_min >= ALERT_ARRET_PROLONGE_MINUTES:
                stop_alerts.append({
                    "lat": pts[i]["lat"], "lon": pts[i]["lon"],
                    "duree_minutes": round(duree_min, 1), "time": pts[i]["time"]
                })
            i = j
        else:
            i += 1

    return {
        "corridors": corridors,
        "positions": pos_out,
        "alerts": {
            "vitesse": speed_alerts,
            "seuil_vitesse_kmh": ALERT_MAX_WORKING_SPEED_KMH,
            "arrets_prolonges": stop_alerts,
            "seuil_arret_minutes": ALERT_ARRET_PROLONGE_MINUTES,
        },
        "stats": {
            "dist_km":    round(total_dist / 1000, 2),
            "area_ha":    round(total_area, 2),
            "avg_speed":  avg_speed,
            "nb_passages": len(corridors),
            "date":       date_str,
        }
    }


# NDVI (tuiles colorisees, config, historique) : extrait dans ndvi_bp.py -- voir
# l'enregistrement du blueprint en tete de fichier.

# ================= EXPORT PDF DE SYNTHESE DE CHANTIER =================
VEHICLE_COLORS_HEX = ['#22c55e','#3b82f6','#f59e0b','#e11d48','#a855f7','#06b6d4','#f97316','#84cc16','#ec4899','#14b8a6']

def _get_catalog_tools():
    """Lit la table catalog_tools (mot-clé outil -> nom d'intervention), même source que index.html."""
    try:
        DB_PATH = 'database.db'
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT keyword, intervention FROM catalog_tools")
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []

def _deduce_intervention_from_tool(tool, catalog):
    """Déduit le nom de l'intervention à partir du mot-clé outil détecté (même logique que le frontend)."""
    if not tool:
        return None
    t_upper = str(tool).upper().strip()
    for entry in catalog:
        kw = str(entry.get("keyword", "")).upper().strip()
        if kw and kw in t_upper:
            return entry.get("intervention")
    return None

def _hex_to_rgb(hexcolor):
    h = hexcolor.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def _color_for_device_index(idx):
    return VEHICLE_COLORS_HEX[idx % len(VEHICLE_COLORS_HEX)]

@app.route("/api/export_pdf")
@login_required
def api_export_pdf():
    """
    Genere un PDF de synthese pour un chantier : legende vehicules (avec nom d'intervention
    deduit de l'outil), schema cartographique (trace des passages, largeur proportionnelle a
    la largeur d'outil reelle, contour des parcelles), stats globales et par vehicule, et
    couverture pour chaque parcelle detectee/indiquee. Ne depend d'aucune tuile/carte internet :
    tout est reconstruit a partir des coordonnees GPS (projection equirectangulaire locale simple).
    Parametres : date (YYYY-MM-DD), device_ids (separes par des virgules),
                 geofence_ids (optionnel, separes par des virgules -- ou geofence_id au singulier)
    """
    import math

    date_str   = request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))
    device_ids = [d for d in request.args.get("device_ids", "").split(",") if d]
    geofence_ids = [g for g in request.args.get("geofence_ids", "").split(",") if g]
    single_gf = request.args.get("geofence_id")
    if single_gf and single_gf not in geofence_ids:
        geofence_ids.append(single_gf)

    if not device_ids:
        return jsonify({"error": "device_ids requis"}), 400

    devices = get_devices()
    catalog_tools = _get_catalog_tools()

    # ── Récupération des données par véhicule ──
    per_vehicle = []
    for idx, dev_id in enumerate(device_ids):
        dev_info = devices.get(int(dev_id)) if str(dev_id).isdigit() else None
        name = dev_info.get("name", f"Véhicule {dev_id}") if dev_info else f"Véhicule {dev_id}"
        color = _color_for_device_index(idx)
        pts = _fetch_and_parse_positions(dev_id, date_str)
        data = {"stats": {}, "positions": [], "corridors": []}
        if pts:
            try:
                data = _build_corridors_response(pts, date_str, math)
            except Exception:
                app.logger.exception("export_pdf: erreur calcul corridors pour device %s", dev_id)

        # Intervention dominante du véhicule sur la journée : outil le plus fréquent parmi
        # les passages, déduit en nom d'intervention via le catalogue outil -> intervention.
        corridors = data.get("corridors", [])
        tools = [c.get("tool") for c in corridors if c.get("tool")]
        dominant_tool = max(set(tools), key=tools.count) if tools else None
        intervention_name = _deduce_intervention_from_tool(dominant_tool, catalog_tools)

        per_vehicle.append({
            "id": dev_id, "name": name, "color": color,
            "stats": data.get("stats", {}), "positions": data.get("positions", []),
            "corridors": corridors, "intervention": intervention_name,
        })

    # ── Géofences (optionnelles) pour le contour + l'analyse de couverture ──
    geofence_list = []
    if geofence_ids:
        raw_geofences = safe_get(f"{TRACCAR_URL}/geofences")
        for gid in geofence_ids:
            gf = next((g for g in raw_geofences if str(g.get("id")) == str(gid)), None) if isinstance(raw_geofences, list) else None
            if not gf:
                continue
            gname = gf.get("name", f"Parcelle {gid}")
            geom = _parse_geofence_wkt(gf.get("area"))
            coords = None
            if geom:
                if geom["type"] == "polygon":
                    coords = geom["coords"]
                else:
                    clat, clon = geom["center"]; r = geom["radius"]
                    coords = []
                    for a in range(0, 360, 15):
                        dlat = (r / 111320.0) * math.cos(math.radians(a))
                        dlon = (r / (111320.0 * max(0.0001, math.cos(math.radians(clat))))) * math.sin(math.radians(a))
                        coords.append([clat + dlat, clon + dlon])
            cov = None
            try:
                cov_result, cov_status = _compute_coverage(gid, date_str, device_ids, 8.0)
                if cov_status == 200 and "coverage_pct" in cov_result:
                    cov = cov_result
            except Exception:
                app.logger.exception("export_pdf: erreur calcul couverture pour geofence %s", gid)
            geofence_list.append({"name": gname, "coords": coords, "coverage": cov})

    try:
        pdf_path = _generate_chantier_pdf(date_str, per_vehicle, geofence_list)
    except Exception as e:
        app.logger.exception("export_pdf: erreur generation PDF")
        return jsonify({"error": f"Erreur génération PDF : {e}"}), 500

    filename = f"synthese_chantier_{date_str}.pdf"
    return send_file(pdf_path, mimetype="application/pdf",
                      as_attachment=True, download_name=filename)


def _generate_chantier_pdf(date_str, per_vehicle, geofence_list):
    def safe(t):
        return str(t if t is not None else '').encode('latin-1', 'replace').decode('latin-1')

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    page_w = 277  # largeur utile en paysage A4 avec marges ~10mm de chaque cote

    # ── En-tete ──
    date_fr = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
    interventions_presentes = [v["intervention"] for v in per_vehicle if v.get("intervention")]
    dominant_intervention = max(set(interventions_presentes), key=interventions_presentes.count) if interventions_presentes else None
    title_parts = ["Synthèse de chantier"]
    if dominant_intervention:
        title_parts.append(dominant_intervention)
    title_parts.append(date_fr)
    # NB: le tiret cadratin "—" n'existe pas en Latin-1 (jeu de caracteres des polices FPDF de
    # base) : encode('latin-1','replace') le transforme silencieusement en "?" dans le PDF.
    # On utilise donc un simple tiret "-" partout dans ce document.
    pdf.set_fill_color(22, 163, 74)  # vert (#16a34a)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 12, safe(" - ".join(title_parts)), ln=1, fill=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # ── Légende véhicules (avec nom d'intervention déduit de l'outil) ──
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, safe("Véhicules :"), ln=1)
    pdf.set_font("Arial", "", 10)
    for v in per_vehicle:
        r, g, b = _hex_to_rgb(v["color"])
        pdf.set_fill_color(r, g, b)
        pdf.rect(pdf.get_x(), pdf.get_y() + 1, 4, 4, style="F")
        pdf.cell(6, 6, "")
        label = v["name"] + (f" - {v['intervention']}" if v.get("intervention") else "")
        pdf.cell(90, 6, safe(label))
    pdf.ln(9)

    map_x, map_y = 10, pdf.get_y()
    map_w, map_h = 165, 150
    stats_x = map_x + map_w + 8

    # ── Schéma cartographique (sans tuile internet : projection locale simple) ──
    all_pts_for_bbox = []
    for gf in geofence_list:
        if gf.get("coords"):
            all_pts_for_bbox.extend(gf["coords"])
    for v in per_vehicle:
        all_pts_for_bbox.extend([[p["lat"], p["lon"]] for p in v["positions"]])

    pdf.set_draw_color(51, 65, 85)
    pdf.rect(map_x, map_y, map_w, map_h)

    if len(all_pts_for_bbox) >= 2:
        lats = [p[0] for p in all_pts_for_bbox]; lons = [p[1] for p in all_pts_for_bbox]
        lat_min, lat_max = min(lats), max(lats)
        lon_min, lon_max = min(lons), max(lons)
        mid_lat = (lat_min + lat_max) / 2.0
        import math
        m_per_deg_lat = 111320.0
        m_per_deg_lon = 111320.0 * max(0.0001, math.cos(math.radians(mid_lat)))
        span_lat_m = max((lat_max - lat_min) * m_per_deg_lat, 1.0)
        span_lon_m = max((lon_max - lon_min) * m_per_deg_lon, 1.0)
        margin_mm = 6
        scale = min((map_w - 2*margin_mm) / span_lon_m, (map_h - 2*margin_mm) / span_lat_m)

        def proj(lat, lon):
            x = map_x + margin_mm + (lon - lon_min) * m_per_deg_lon * scale
            y = map_y + map_h - margin_mm - (lat - lat_min) * m_per_deg_lat * scale
            return x, y

        # Contour de chaque parcelle (si fournie(s)), en pointillé jaune
        for gf in geofence_list:
            if not gf.get("coords"):
                continue
            pdf.set_draw_color(234, 179, 8)
            pdf.set_line_width(0.6)
            pts_px = [proj(la, lo) for la, lo in gf["coords"]]
            for i in range(len(pts_px)):
                x1, y1 = pts_px[i]; x2, y2 = pts_px[(i+1) % len(pts_px)]
                pdf.line(x1, y1, x2, y2)

        # Trace de chaque véhicule : un segment de centre-ligne par quadrilatère de corridor,
        # avec l'épaisseur de trait proportionnelle à la LARGEUR RÉELLE de l'outil pour ce
        # passage (et non une valeur devinée) — le centre-ligne est le milieu gauche/droite
        # de chaque quadrilatère, déjà calculé côté serveur avec un cap unique par segment.
        for v in per_vehicle:
            r, g, b = _hex_to_rgb(v["color"])
            pdf.set_draw_color(r, g, b)
            for corridor in v.get("corridors", []):
                width_m = corridor.get("width", 3.0)
                pdf.set_line_width(max(0.5, width_m * scale))
                for q in corridor.get("quads", []):
                    l_k, l_k1, r_k1, r_k = q
                    mid1 = ((l_k[0] + r_k[0]) / 2.0, (l_k[1] + r_k[1]) / 2.0)
                    mid2 = ((l_k1[0] + r_k1[0]) / 2.0, (l_k1[1] + r_k1[1]) / 2.0)
                    x1, y1 = proj(*mid1); x2, y2 = proj(*mid2)
                    pdf.line(x1, y1, x2, y2)
        pdf.set_line_width(0.2)
    else:
        pdf.set_xy(map_x, map_y + map_h/2 - 4)
        pdf.set_font("Arial", "I", 10)
        pdf.cell(map_w, 8, safe("Pas assez de positions GPS pour tracer un schéma"), align="C")

    # ── Colonne de droite : stats ──
    pdf.set_xy(stats_x, map_y)
    col_w = page_w - (stats_x - 10) - 10

    pdf.set_font("Arial", "B", 11)
    pdf.set_fill_color(30, 41, 59); pdf.set_text_color(255, 255, 255)
    pdf.cell(col_w, 8, safe("  Statistiques globales"), ln=1, fill=True)
    pdf.set_text_color(0, 0, 0); pdf.set_x(stats_x)

    total_area = sum(v["stats"].get("area_ha", 0) or 0 for v in per_vehicle)
    total_dist = sum(v["stats"].get("dist_km", 0) or 0 for v in per_vehicle)

    pdf.set_font("Arial", "", 10)
    rows = [
        ("Surface travaillée", f"{round(total_area,2)} ha"),
        ("Distance parcourue", f"{round(total_dist,2)} km"),
    ]
    fill = False
    for label, val in rows:
        pdf.set_x(stats_x)
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(col_w*0.6, 7, safe(f"  {label}"), border=1, fill=fill)
        pdf.cell(col_w*0.4, 7, safe(val), border=1, align="R", fill=fill)
        pdf.ln()

    for gf in geofence_list:
        cov = gf.get("coverage")
        if not cov or cov.get("coverage_pct") is None or "message" in cov:
            continue
        pdf.ln(3)
        pdf.set_x(stats_x)
        pdf.set_font("Arial", "B", 11)
        pdf.set_fill_color(29, 78, 216); pdf.set_text_color(255, 255, 255)
        pdf.cell(col_w, 8, safe(f"  Couverture - {gf['name']}"), ln=1, fill=True)
        pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 10)
        cov_rows = [
            ("Taux de couverture", f"{cov['coverage_pct']} %"),
            ("Surface parcelle", f"{cov.get('geofence_area_ha','-')} ha"),
            ("Surface non couverte", f"{cov.get('uncovered_area_ha','-')} ha"),
        ]
        for label, val in cov_rows:
            pdf.set_x(stats_x)
            pdf.set_fill_color(241, 245, 249)
            pdf.cell(col_w*0.6, 7, safe(f"  {label}"), border=1)
            pdf.cell(col_w*0.4, 7, safe(val), border=1, align="R")
            pdf.ln()

    pdf.ln(4)
    pdf.set_x(stats_x)
    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(30, 41, 59); pdf.set_text_color(255, 255, 255)
    pdf.cell(col_w, 7, safe("  Détail par véhicule"), ln=1, fill=True)
    pdf.set_text_color(0, 0, 0)

    pdf.set_font("Arial", "B", 9)
    headers = ["Véhicule", "Intervention", "ha", "km", "km/h"]
    widths  = [col_w*0.28, col_w*0.28, col_w*0.16, col_w*0.16, col_w*0.12]
    pdf.set_x(stats_x)
    pdf.set_fill_color(226, 232, 240)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 7, safe(h), border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for v in per_vehicle:
        s = v["stats"]
        pdf.set_x(stats_x)
        pdf.cell(widths[0], 7, safe(v["name"]), border=1)
        pdf.cell(widths[1], 7, safe(v.get("intervention") or "-"), border=1)
        pdf.cell(widths[2], 7, safe(s.get("area_ha", "-")), border=1, align="C")
        pdf.cell(widths[3], 7, safe(s.get("dist_km", "-")), border=1, align="C")
        pdf.cell(widths[4], 7, safe(s.get("avg_speed", "-")), border=1, align="C")
        pdf.ln()

    # Désactive le saut de page automatique juste pour le pied de page : sans ça, la position
    # set_y(-15) (à 15mm du bas) combinée à la marge de saut de page (12mm) peut déclencher
    # une page supplémentaire rien que pour afficher cette ligne de texte, laissant une page
    # quasi vide à la fin du PDF.
    pdf.set_auto_page_break(auto=False)
    pdf.set_y(-15)
    pdf.set_font("Arial", "I", 8)
    pdf.cell(0, 5, safe(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - Dashboard Agricole v{DASHBOARD_VERSION}"), align="R")

    os.makedirs('exports', exist_ok=True)
    path = os.path.join('exports', f"synthese_chantier_{date_str.replace('-','')}.pdf")
    pdf.output(path)
    return path


# ================= EXPORT GEOJSON / SHAPEFILE (SIG) =================
def _collect_export_features(date_str, device_ids):
    """
    Récupère, pour chaque véhicule, les polygones de passage (corridor.polygon) avec leurs
    attributs, prêts à être exportés en GeoJSON ou Shapefile.
    Retourne une liste de {"coords": [[lon,lat],...], "attrs": {...}}.
    """
    import math
    devices = get_devices()
    catalog_tools = _get_catalog_tools()
    features = []
    for idx, dev_id in enumerate(device_ids):
        dev_info = devices.get(int(dev_id)) if str(dev_id).isdigit() else None
        name = dev_info.get("name", f"Véhicule {dev_id}") if dev_info else f"Véhicule {dev_id}"
        pts = _fetch_and_parse_positions(dev_id, date_str)
        if not pts:
            continue
        try:
            data = _build_corridors_response(pts, date_str, math)
        except Exception:
            continue
        for i, corridor in enumerate(data.get("corridors", [])):
            # GeoJSON/Shapefile utilisent l'ordre (lon, lat), l'inverse de notre convention interne
            coords = [[lon, lat] for lat, lon in corridor["polygon"]]
            if coords and coords[0] != coords[-1]:
                coords.append(coords[0])
            tool = corridor.get("tool", "")
            intervention = _deduce_intervention_from_tool(tool, catalog_tools)
            features.append({
                "coords": coords,
                "attrs": {
                    "vehicule": name,
                    "date": date_str,
                    "passage": i + 1,
                    "outil": tool or "",
                    "intervention": intervention or "",
                    "largeur_m": corridor.get("width", 0),
                    "surface_ha": corridor.get("area_ha", 0),
                    "distance_m": corridor.get("dist_m", 0),
                }
            })
    return features


@app.route("/api/export_geojson")
@login_required
def api_export_geojson():
    """Exporte les passages travaillés (polygones) au format GeoJSON standard (EPSG:4326)."""
    date_str   = request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))
    device_ids = [d for d in request.args.get("device_ids", "").split(",") if d]
    if not device_ids:
        return jsonify({"error": "device_ids requis"}), 400

    features = _collect_export_features(date_str, device_ids)
    geojson = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "geometry": {"type": "Polygon", "coordinates": [f["coords"]]},
                "properties": f["attrs"],
            }
            for f in features if len(f["coords"]) >= 4
        ]
    }
    buf = io.BytesIO(json.dumps(geojson, ensure_ascii=False, indent=2).encode("utf-8"))
    return send_file(buf, mimetype="application/geo+json", as_attachment=True,
                      download_name=f"corridors_{date_str}.geojson")


def _dbf_field(name, ftype, length, decimals=0):
    name_b = name.encode("ascii", "replace")[:10].ljust(11, b"\x00")
    return name_b + ftype.encode("ascii") + b"\x00\x00\x00\x00" + bytes([length]) + bytes([decimals]) + b"\x00" * 14

def _write_dbf(records, fields):
    """
    Ecrit un fichier .dbf (dBase III) minimal, en pur Python (aucune dépendance externe).
    fields: liste de (nom, type 'C'/'N', longueur, decimales)
    records: liste de dicts {nom: valeur}
    """
    num_records = len(records)
    header_size = 32 + 32 * len(fields) + 1
    record_size = 1 + sum(f[2] for f in fields)  # 1 = flag de suppression

    buf = io.BytesIO()
    today = datetime.utcnow()
    buf.write(bytes([0x03]))
    buf.write(bytes([today.year - 1900, today.month, today.day]))
    buf.write(struct.pack('<i', num_records))
    buf.write(struct.pack('<h', header_size))
    buf.write(struct.pack('<h', record_size))
    buf.write(b"\x00" * 2)   # reserved
    buf.write(bytes([0]))    # incomplete transaction
    buf.write(bytes([0]))    # encryption
    buf.write(b"\x00" * 12)  # reserved multi-user
    buf.write(bytes([0]))    # MDX flag
    buf.write(bytes([0]))    # language driver
    buf.write(b"\x00" * 2)   # reserved

    for name, ftype, length, decimals in fields:
        buf.write(_dbf_field(name, ftype, length, decimals))
    buf.write(b"\x0D")  # terminateur d'en-tête

    for rec in records:
        buf.write(b" ")  # non supprimé
        for name, ftype, length, decimals in fields:
            val = rec.get(name, "")
            if ftype == "N":
                s = f"{float(val):.{decimals}f}" if decimals else str(int(float(val)))
                s = s[:length].rjust(length)
            else:
                s = str(val).encode("latin-1", "replace").decode("latin-1")[:length].ljust(length)
            buf.write(s.encode("latin-1", "replace"))
    buf.write(b"\x1A")  # fin de fichier
    return buf.getvalue()

def _write_shp_shx(polygons):
    """
    Ecrit les fichiers .shp et .shx (polygones, EPSG:4326) en pur Python, sans dépendance
    externe (pas de pyshp/GDAL nécessaire côté serveur). polygons: liste d'anneaux
    [[lon,lat], ...] fermés (premier point = dernier point).
    """
    records = []
    for coords in polygons:
        if coords[0] != coords[-1]:
            coords = coords + [coords[0]]
        xs = [c[0] for c in coords]; ys = [c[1] for c in coords]
        box = (min(xs), min(ys), max(xs), max(ys))
        content = struct.pack('<i', 5)  # shape type = polygon
        content += struct.pack('<4d', *box)
        content += struct.pack('<i', 1)           # numParts (un seul anneau, pas de trou)
        content += struct.pack('<i', len(coords))  # numPoints
        content += struct.pack('<i', 0)           # parts[0] = index de depart
        for x, y in coords:
            content += struct.pack('<2d', x, y)
        records.append(content)

    all_x = [c[0] for poly in polygons for c in poly]
    all_y = [c[1] for poly in polygons for c in poly]
    global_box = (min(all_x), min(all_y), max(all_x), max(all_y), 0, 0, 0, 0)

    def write_header(buf, file_len_words):
        buf.write(struct.pack('>i', 9994))
        buf.write(b'\x00' * 20)
        buf.write(struct.pack('>i', file_len_words))
        buf.write(struct.pack('<i', 1000))
        buf.write(struct.pack('<i', 5))
        buf.write(struct.pack('<8d', *global_box))

    shp = io.BytesIO(); shx = io.BytesIO()
    file_len_words = (100 + sum(8 + len(r) for r in records)) // 2
    shx_len_words = (100 + 8 * len(records)) // 2
    write_header(shp, file_len_words)
    write_header(shx, shx_len_words)

    offset_words = 50
    for i, content in enumerate(records, start=1):
        rec_len_words = len(content) // 2
        shp.write(struct.pack('>i', i)); shp.write(struct.pack('>i', rec_len_words)); shp.write(content)
        shx.write(struct.pack('>i', offset_words)); shx.write(struct.pack('>i', rec_len_words))
        offset_words += 4 + rec_len_words

    return shp.getvalue(), shx.getvalue()

_WGS84_PRJ = (
    'GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137.0,298.257223563]],'
    'PRIMEM["Greenwich",0.0],UNIT["Degree",0.0174532925199433]]'
)

@app.route("/api/export_shapefile")
@login_required
def api_export_shapefile():
    """
    Exporte les passages travaillés au format Shapefile ESRI (.shp/.shx/.dbf/.prj zippés),
    lisible par QGIS et la plupart des logiciels agro. Ecrit en pur Python (aucune dépendance
    type pyshp/GDAL requise côté serveur).
    """
    date_str   = request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))
    device_ids = [d for d in request.args.get("device_ids", "").split(",") if d]
    if not device_ids:
        return jsonify({"error": "device_ids requis"}), 400

    features = [f for f in _collect_export_features(date_str, device_ids) if len(f["coords"]) >= 4]
    if not features:
        return jsonify({"error": "Aucun passage à exporter pour cette sélection"}), 404

    polygons = [f["coords"] for f in features]
    fields = [
        ("vehicule", "C", 40, 0),
        ("date", "C", 10, 0),
        ("passage", "N", 6, 0),
        ("outil", "C", 30, 0),
        ("interv", "C", 30, 0),  # "intervention" tronqué : les noms de champs DBF sont limités à 10 caractères
        ("largeur_m", "N", 8, 2),
        ("surface_ha", "N", 10, 4),
        ("distance_m", "N", 10, 1),
    ]
    dbf_records = [{
        "vehicule": f["attrs"]["vehicule"], "date": f["attrs"]["date"], "passage": f["attrs"]["passage"],
        "outil": f["attrs"].get("outil", ""), "interv": f["attrs"].get("intervention", ""),
        "largeur_m": f["attrs"]["largeur_m"], "surface_ha": f["attrs"]["surface_ha"], "distance_m": f["attrs"]["distance_m"],
    } for f in features]

    shp_bytes, shx_bytes = _write_shp_shx(polygons)
    dbf_bytes = _write_dbf(dbf_records, fields)

    zbuf = io.BytesIO()
    base = f"corridors_{date_str}"
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{base}.shp", shp_bytes)
        zf.writestr(f"{base}.shx", shx_bytes)
        zf.writestr(f"{base}.dbf", dbf_bytes)
        zf.writestr(f"{base}.prj", _WGS84_PRJ)
    zbuf.seek(0)
    return send_file(zbuf, mimetype="application/zip", as_attachment=True,
                      download_name=f"{base}_shapefile.zip")


# ================= RESUME HEBDOMADAIRE / MENSUEL PAR VEHICULE (EXCEL) =================
def _daterange_str(start_str, end_str):
    """Liste de dates (YYYY-MM-DD) entre deux bornes incluses."""
    d0 = datetime.strptime(start_str, "%Y-%m-%d")
    d1 = datetime.strptime(end_str, "%Y-%m-%d")
    if d1 < d0:
        d0, d1 = d1, d0
    days = []
    d = d0
    while d <= d1:
        days.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)
    return days

def _compute_stats_for_device_day(dev_id, date_str):
    """Calcule les stats de corridors pour un véhicule et une date donnée (ou None si rien)."""
    import math
    pts = _fetch_and_parse_positions(dev_id, date_str)
    if not pts:
        return None
    try:
        data = _build_corridors_response(pts, date_str, math)
    except Exception:
        return None
    stats = data.get("stats", {})
    if not stats.get("nb_passages"):
        return None
    return stats

@app.route("/api/export_summary_excel")
@login_required
def api_export_summary_excel():
    """
    Génère un classeur Excel (résumé hebdomadaire/mensuel par véhicule) sur une période donnée :
    - Feuille "Détail journalier" : une ligne par (véhicule, jour) avec activité.
    - Feuille "Résumé par véhicule" : totaux/moyennes agrégés sur toute la période.
    Paramètres : start_date, end_date (YYYY-MM-DD), device_ids (séparés par des virgules)
    """
    start_date = request.args.get("start_date")
    end_date   = request.args.get("end_date")
    device_ids = [d for d in request.args.get("device_ids", "").split(",") if d]

    if not start_date or not end_date or not device_ids:
        return jsonify({"error": "start_date, end_date et device_ids requis"}), 400
    try:
        dates = _daterange_str(start_date, end_date)
    except ValueError:
        return jsonify({"error": "Dates invalides, format YYYY-MM-DD"}), 400

    MAX_JOBS = 400  # garde-fou : evite une periode demesurement longue x trop de vehicules
    if len(dates) * len(device_ids) > MAX_JOBS:
        return jsonify({"error": f"Trop de combinaisons véhicule × jour ({len(dates)*len(device_ids)} > {MAX_JOBS}). Réduisez la période ou le nombre de véhicules."}), 400

    devices = get_devices()
    dev_names = {}
    for dev_id in device_ids:
        info = devices.get(int(dev_id)) if str(dev_id).isdigit() else None
        dev_names[dev_id] = info.get("name", f"Véhicule {dev_id}") if info else f"Véhicule {dev_id}"

    # ── Calcul parallèle (véhicule x jour), même principe que get_data_parallel ──
    results = {}  # (dev_id, date_str) -> stats ou None
    tasks = [(dev_id, d) for dev_id in device_ids for d in dates]
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(_compute_stats_for_device_day, dev_id, d): (dev_id, d) for dev_id, d in tasks}
        for fut in as_completed(futures):
            key = futures[fut]
            try:
                results[key] = fut.result()
            except Exception:
                results[key] = None

    # ── Construction du classeur ──
    wb = Workbook()
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill  = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    total_font  = Font(bold=True)

    # -- Feuille 1 : détail journalier --
    ws1 = wb.active
    ws1.title = "Détail journalier"
    headers1 = ["Date", "Véhicule", "Surface (ha)", "Distance (km)", "Vitesse moy. (km/h)", "Passages"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center")

    for d in dates:
        for dev_id in device_ids:
            stats = results.get((dev_id, d))
            if not stats:
                continue
            ws1.append([
                d, dev_names[dev_id],
                stats.get("area_ha", 0), stats.get("dist_km", 0),
                stats.get("avg_speed", 0), stats.get("nb_passages", 0)
            ])
    for i, w in enumerate([12, 26, 14, 14, 18, 12], start=1):
        ws1.column_dimensions[chr(64+i)].width = w

    # -- Feuille 2 : résumé par véhicule --
    ws2 = wb.create_sheet("Résumé par véhicule")
    headers2 = ["Véhicule", "Jours actifs", "Surface totale (ha)", "Distance totale (km)",
                "Passages totaux", "Surface moy./jour actif (ha)", "Vitesse moy. pondérée (km/h)"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center")

    grand_total_area, grand_total_dist, grand_total_passages, grand_total_days = 0, 0, 0, 0
    for dev_id in device_ids:
        jours_actifs, total_area, total_dist, total_passages = 0, 0.0, 0.0, 0
        speed_weighted_sum, speed_weight = 0.0, 0.0
        for d in dates:
            stats = results.get((dev_id, d))
            if not stats:
                continue
            jours_actifs += 1
            total_area += stats.get("area_ha", 0) or 0
            total_dist += stats.get("dist_km", 0) or 0
            total_passages += stats.get("nb_passages", 0) or 0
            if stats.get("avg_speed") and stats.get("dist_km"):
                speed_weighted_sum += stats["avg_speed"] * stats["dist_km"]
                speed_weight += stats["dist_km"]

        avg_ha_per_day = round(total_area / jours_actifs, 2) if jours_actifs else 0
        avg_speed_weighted = round(speed_weighted_sum / speed_weight, 1) if speed_weight else 0

        ws2.append([
            dev_names[dev_id], jours_actifs, round(total_area, 2), round(total_dist, 2),
            total_passages, avg_ha_per_day, avg_speed_weighted
        ])
        grand_total_area += total_area; grand_total_dist += total_dist
        grand_total_passages += total_passages; grand_total_days += jours_actifs

    ws2.append(["TOTAL", grand_total_days, round(grand_total_area, 2), round(grand_total_dist, 2),
                grand_total_passages, "", ""])
    for cell in ws2[ws2.max_row]:
        cell.fill = total_fill; cell.font = total_font

    for i, w in enumerate([26, 13, 18, 18, 15, 24, 22], start=1):
        ws2.column_dimensions[chr(64+i)].width = w

    ws2.append([])
    ws2.append([f"Période analysée : {start_date} au {end_date}"])
    ws2.append([f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — Dashboard Agricole v{DASHBOARD_VERSION}"])

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True,
                      download_name=f"resume_{start_date}_au_{end_date}.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/chantier")
@login_required
def chantier():
    """Page carte de travail du jour."""
    return render_template("chantier.html")


# Initialisation au chargement du module (plus sous condition "if __name__ == '__main__'")
# -- dashboard.py doit toujours être *importé*, jamais exécuté directement avec
# "python dashboard.py" : les blueprints (ndvi_bp.py notamment) font "import dashboard" pour
# accéder à son état partagé (TRACCAR_URL...), ce qui ne fonctionne correctement que si
# "dashboard" est le nom sous lequel Python a chargé ce fichier. Lancer le serveur via
# "python run.py" (voir ce fichier) au lieu de "python dashboard.py" directement.
init_db()
backup_database()
threading.Thread(target=backup_scheduler, daemon=True).start()
