"""
Blueprint des exports du rapport de chantier (Excel et PDF) : liste des interventions
GPS-detectees filtrees par vehicule/parcelle/periode, telechargeable directement.

Extrait de dashboard.py. Depend de build_data() et apply_filters() (references via
"import dashboard", meme motif que ndvi_bp.py/cahier_bp.py -- apply_filters reste dans le
noyau car aussi utilisee par la route /data principale du dashboard).
"""
import os
import io
import re
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF
from flask import Blueprint, request, jsonify, send_file, redirect, url_for, session as flask_session

import dashboard

chantier_export_bp = Blueprint("chantier_export", __name__)


@chantier_export_bp.before_request
def _require_login():
    """
    Meme authentification que le reste de l'application -- avec la meme distinction que le
    decorateur login_required d'origine : /export_excel et /export_pdf ne sont PAS sous /api/
    (telechargements directs, pas des appels JSON), donc une session expiree doit rediriger
    vers la page de connexion plutot que de renvoyer une erreur JSON brute dans le navigateur.
    """
    if not flask_session.get("logged_in"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Non authentifie", "redirect": "/login"}), 401
        return redirect(url_for("login"))


@chantier_export_bp.route("/export_excel")
def export_excel():
    raw = dashboard.build_data()
    event_data = dashboard.apply_filters(raw["events"], request.args.get("vehicle",""), request.args.get("geofence",""), request.args.get("start"), request.args.get("end"))
    wb = Workbook()
    ws = wb.active
    ws.append(["Vehicle","Geofence","Type","Date","Duration","Surf. Parcelle","Surf. Travaillee","Tool","Width"])
    for d in event_data:
        ws.append([d["vehicle"], d["geofence"], d["type"], d["date_fr"], d.get("duration","-"), d.get("field","-"), d.get("appliedArea","-"), d["tool"], d["width"]])
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name="traccar_v7.14.xlsx")

@chantier_export_bp.route("/export_pdf")
def export_pdf():
    raw = dashboard.build_data()
    event_data = dashboard.apply_filters(raw["events"], request.args.get("vehicle",""), request.args.get("geofence",""), request.args.get("start"), request.args.get("end"))
    event_data.sort(key=lambda x: x["geofence"])
    
    geo_totals = {}
    for d in event_data:
        if d["type"] == "Sortie" and d.get("duration") and d["duration"] != "-":
            minutes = 0
            match_h = re.search(r'(\d+)h', d["duration"])
            match_m = re.search(r'(\d+)m', d["duration"])
            if match_h: minutes += int(match_h.group(1)) * 60
            if match_m: minutes += int(match_m.group(1))
            geo_totals[d["geofence"]] = geo_totals.get(d["geofence"], 0) + minutes

    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Rapport Chantiers Traccar - v7.14", ln=1, align="C")
    
    headers = ["Parcelle", "Surf. Parc", "Surf. Trav", "Date", "Type", "Durée", "Véhicule", "Outil", "Largeur"]
    page_width = pdf.w - 2 * pdf.l_margin
    widths = [
        page_width * 0.14, 
        page_width * 0.08, 
        page_width * 0.08, 
        page_width * 0.14, 
        page_width * 0.07, 
        page_width * 0.09, 
        page_width * 0.14, 
        page_width * 0.18, 
        page_width * 0.08  
    ]
    
    pdf.set_font("Arial", "B", 9)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 10, h.encode('latin-1', 'replace').decode('latin-1'), 1, align="C")
    pdf.ln()
    
    pdf.set_font("Arial", "", 9)
    current_geo = None
    for d in event_data:
        if d["geofence"] != current_geo:
            current_geo = d["geofence"]
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(220, 220, 220)
            
            total_min = geo_totals.get(current_geo, 0)
            if total_min > 0:
                h = total_min // 60
                m = total_min % 60
                total_str = f"Temps total : {h}h{m:02d}m" if h > 0 else f"Temps total : {m}m"
            else:
                total_str = "Temps total : 0m"

            title_txt = f" Parcelle : {current_geo}"
            pdf.cell(page_width * 0.6, 8, title_txt.encode('latin-1', 'replace').decode('latin-1'), 1, 0, fill=True)
            pdf.cell(page_width * 0.4, 8, f"{total_str} ", 1, 1, align="R", fill=True)
            pdf.set_font("Arial", "", 9)
            
        pdf.set_fill_color(*(212, 247, 212) if d["type"] == "Entrée" else (255, 214, 214))
        
        row = [
            d["geofence"], 
            d.get("field", "-"), 
            d.get("appliedArea", "-"), 
            d["date_fr"], 
            d["type"], 
            d.get("duration","-"), 
            d["vehicle"], 
            d["tool"],
            d["width"]
        ]
        for i, v in enumerate(row):
            val_txt = str(v if v else "-")
            pdf.cell(widths[i], 8, val_txt.encode('latin-1', 'replace').decode('latin-1'), 1, fill=True)
        pdf.ln()
        
    os.makedirs("exports", exist_ok=True)
    path = os.path.join("exports", "report.pdf")
    pdf.output(path)
    return send_file(path, as_attachment=True)


def _ics_echapper(texte):
    """Echappe les caracteres speciaux du format iCalendar (RFC 5545) : les virgules,
    points-virgules et retours a la ligne ont un sens special dans ce format."""
    if texte is None:
        return ""
    return (str(texte)
            .replace("\\", "\\\\")
            .replace(",", "\\,")
            .replace(";", "\\;")
            .replace("\n", "\\n"))


@chantier_export_bp.route("/export_ics")
def export_ics():
    """
    Exporte les interventions VALIDEES (celles enregistrees dans le Carnet, memes donnees
    que celles affichees dans la modale Calendrier -- table SQLite "interventions") au
    format iCalendar (.ics), importable dans Google Calendar / Outlook / Apple Calendar.

    Ne pas confondre avec les "events" de dashboard.build_data() : ce sont de simples
    passages GPS detectes par Traccar (entrees/sorties de geofence), qui n'ont pas
    forcement ete transformes en intervention validee dans le Carnet. Le calendrier de
    l'application affiche les interventions VALIDEES, donc c'est cette meme source qu'il
    faut exporter ici pour que le contenu du .ics corresponde a ce que l'utilisateur voit
    dans la modale Calendrier.

    Parametres optionnels : start, end (YYYY-MM-DD) pour filtrer par date d'intervention.
    """
    import sqlite3 as _sqlite3
    import json as _json

    start = request.args.get("start")  # format YYYY-MM-DD
    end = request.args.get("end")

    with _sqlite3.connect("database.db") as conn:
        conn.row_factory = _sqlite3.Row
        cur = conn.cursor()

        query = """
            SELECT device_id, geofence_id, exit_time, vehicle_name, tool_detected,
                   intervention_type, products, applied_area, duration_min, rendement,
                   sous_parcelle_id
            FROM interventions
        """
        conds, params = [], []
        if start:
            conds.append("exit_time >= ?"); params.append(start)
        if end:
            conds.append("exit_time <= ?"); params.append(end + "T23:59:59")
        if conds:
            query += " WHERE " + " AND ".join(conds)
        query += " ORDER BY exit_time ASC"
        cur.execute(query, params)
        interventions = [dict(r) for r in cur.fetchall()]

        cur.execute("SELECT geofence_id, nom_parcelle FROM parcelles")
        noms_parcelles = {str(r["geofence_id"]): r["nom_parcelle"] for r in cur.fetchall() if r["nom_parcelle"]}

        dashboard._ensure_sous_parcelles_table()
        sous_parcelles_info = dashboard._get_sous_parcelles_info(conn)

    # Complement pour les parcelles sans nom personnalise en base : nom Traccar en direct
    # (meme source que CACHE_GEOFENCES cote client), avec repli generique en dernier recours.
    try:
        geofences_named = dashboard.build_data().get("geofences", {})
    except Exception:
        geofences_named = {}

    def _nom_parcelle(geo_id, sp_id):
        geo_id_str = str(geo_id)
        base = noms_parcelles.get(geo_id_str) or geofences_named.get(geo_id_str, {}).get("name") or f"Parcelle {geo_id}"
        sp = sous_parcelles_info.get(sp_id) if sp_id else None
        return f"{base} — {sp['nom']}" if sp else base

    lignes = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Dashboard Agricole//Export interventions//FR",
        "CALSCALE:GREGORIAN",
    ]

    maintenant_utc = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    for i, iv in enumerate(interventions):
        date_brute = iv.get("exit_time") or ""
        try:
            # Meme pattern de parsing tolerant que format_date_fr : tronquer aux 19
            # premiers caracteres plutot que d'exiger un format ISO strict.
            dt = datetime.strptime(date_brute[:19], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            continue  # date illisible : on ignore cette ligne plutot que produire un .ics invalide

        # duration_min est deja un entier en minutes (contrairement a l'ancien champ texte
        # "2h30m" des events Traccar bruts) -- pas de parsing regex necessaire ici.
        duree_min = iv.get("duration_min")
        dt_fin = dt + timedelta(minutes=duree_min if duree_min else 60)

        nom_parc = _nom_parcelle(iv["geofence_id"], iv.get("sous_parcelle_id"))
        vehicule = iv.get("vehicle_name") or "Saisie manuelle"
        titre = f"{iv.get('intervention_type') or 'Intervention'} - {vehicule} sur {nom_parc}"

        description_parties = [f"Vehicule : {vehicule}"]
        if iv.get("tool_detected"):
            description_parties.append(f"Outil : {iv['tool_detected']}")
        if iv.get("applied_area"):
            description_parties.append(f"Surface travaillee : {iv['applied_area']} ha")
        if iv.get("rendement"):
            description_parties.append(f"Rendement : {iv['rendement']} t/ha")
        try:
            produits = _json.loads(iv.get("products") or "[]")
        except Exception:
            produits = []
        if produits:
            noms_produits = ", ".join(p.get("name", "") for p in produits if p.get("name"))
            if noms_produits:
                description_parties.append(f"Produits : {noms_produits}")
        description = "\\n".join(_ics_echapper(p) for p in description_parties)

        lignes += [
            "BEGIN:VEVENT",
            f"UID:dashboard-agricole-{i}-{dt.strftime('%Y%m%dT%H%M%S')}@traccar-dashboard",
            f"DTSTAMP:{maintenant_utc}",
            f"DTSTART:{dt.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{dt_fin.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{_ics_echapper(titre)}",
            f"DESCRIPTION:{description}",
            f"LOCATION:{_ics_echapper(nom_parc)}",
            "END:VEVENT",
        ]

    lignes.append("END:VCALENDAR")
    # RFC 5545 impose des fins de ligne CRLF
    contenu = "\r\n".join(lignes)

    return send_file(
        io.BytesIO(contenu.encode("utf-8")),
        as_attachment=True,
        download_name="interventions_traccar.ics",
        mimetype="text/calendar",
    )
